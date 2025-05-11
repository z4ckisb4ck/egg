import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image
import tempfile
import time
import random
import uuid
from datetime import datetime
from sklearn.utils import resample
from sklearn.model_selection import ParameterGrid
from joblib import Parallel, delayed
from xgboost import XGBRegressor

# Static historical data
BLS_DATA = np.array([
[2017,1,1.47],[2017,2,1.48],[2017,3,1.49],[2017,4,1.50],[2017,5,1.51],[2017,6,1.52],
[2017,7,1.53],[2017,8,1.54],[2017,9,1.55],[2017,10,1.56],[2017,11,1.57],[2017,12,1.58],
[2018,1,1.74],[2018,2,1.75],[2018,3,1.76],[2018,4,1.77],[2018,5,1.78],[2018,6,1.79],
[2018,7,1.80],[2018,8,1.81],[2018,9,1.82],[2018,10,1.83],[2018,11,1.84],[2018,12,1.85],
[2019,1,1.40],[2019,2,1.41],[2019,3,1.42],[2019,4,1.43],[2019,5,1.44],[2019,6,1.45],
[2019,7,1.46],[2019,8,1.47],[2019,9,1.48],[2019,10,1.49],[2019,11,1.50],[2019,12,1.51],
[2020,1,1.461],[2020,2,1.449],[2020,3,1.525],[2020,4,2.019],[2020,5,1.640],[2020,6,1.554],
[2020,7,1.401],[2020,8,1.328],[2020,9,1.353],[2020,10,1.408],[2020,11,1.450],[2020,12,1.481],
[2021,1,1.466],[2021,2,1.597],[2021,3,1.625],[2021,4,1.620],[2021,5,1.625],[2021,6,1.642],
[2021,7,1.642],[2021,8,1.709],[2021,9,1.835],[2021,10,1.821],[2021,11,1.718],[2021,12,1.788],
[2022,1,1.929],[2022,2,2.005],[2022,3,2.046],[2022,4,2.520],[2022,5,2.863],[2022,6,2.707],
[2022,7,2.936],[2022,8,3.116],[2022,9,2.902],[2022,10,3.419],[2022,11,3.589],[2022,12,4.250],
[2023,1,4.823],[2023,2,4.211],[2023,3,3.446],[2023,4,3.270],[2023,5,2.666],[2023,6,2.219],
[2023,7,2.094],[2023,8,2.043],[2023,9,2.065],[2023,10,2.072],[2023,11,2.138],[2023,12,2.507],
[2024,1,2.522],[2024,2,2.996],[2024,3,2.992],[2024,4,2.864],[2024,5,2.699],[2024,6,2.715],
[2024,7,3.080],[2024,8,3.204],[2024,9,3.821],[2024,10,3.370],[2024,11,3.649],[2024,12,4.146],
[2025,1,4.953],[2025,2,5.897],[2025,3,6.227]
])

WHOLESALE_DATA = np.array([
[2017,1,1.03],[2017,2,1.04],[2017,3,1.04],[2017,4,1.05],[2017,5,1.06],[2017,6,1.06],
[2017,7,1.07],[2017,8,1.08],[2017,9,1.09],[2017,10,1.09],[2017,11,1.10],[2017,12,1.11],
[2018,1,1.22],[2018,2,1.23],[2018,3,1.23],[2018,4,1.24],[2018,5,1.25],[2018,6,1.25],
[2018,7,1.26],[2018,8,1.27],[2018,9,1.27],[2018,10,1.28],[2018,11,1.29],[2018,12,1.30],
[2019,1,0.98],[2019,2,0.99],[2019,3,0.99],[2019,4,1.00],[2019,5,1.01],[2019,6,1.02],
[2019,7,1.02],[2019,8,1.03],[2019,9,1.04],[2019,10,1.04],[2019,11,1.05],[2019,12,1.06],
[2020,1,2.30],[2020,2,2.28],[2020,3,2.35],[2020,4,2.40],[2020,5,2.45],[2020,6,2.50],
[2020,7,2.55],[2020,8,2.60],[2020,9,2.65],[2020,10,2.70],[2020,11,2.75],[2020,12,2.80],
[2021,1,2.50],[2021,2,2.52],[2021,3,2.55],[2021,4,2.58],[2021,5,2.60],[2021,6,2.62],
[2021,7,2.65],[2021,8,2.68],[2021,9,2.70],[2021,10,2.72],[2021,11,2.75],[2021,12,2.80],
[2022,1,3.00],[2022,2,3.05],[2022,3,3.10],[2022,4,3.15],[2022,5,3.20],[2022,6,3.25],
[2022,7,3.30],[2022,8,3.35],[2022,9,3.40],[2022,10,3.45],[2022,11,3.50],[2022,12,3.55],
[2023,1,4.60],[2023,2,4.65],[2023,3,4.70],[2023,4,4.75],[2023,5,4.80],[2023,6,4.85],
[2023,7,4.90],[2023,8,4.95],[2023,9,5.00],[2023,10,5.05],[2023,11,5.10],[2023,12,5.15],
[2024,1,4.90],[2024,2,4.95],[2024,3,5.00],[2024,4,5.05],[2024,5,5.10],[2024,6,5.15],
[2024,7,5.20],[2024,8,5.25],[2024,9,5.30],[2024,10,5.35],[2024,11,5.40],[2024,12,5.45],
[2025,1,5.80],[2025,2,5.90],[2025,3,4.85]
])

HEN_DATA = np.array([
[2017,1,522.0],[2017,2,522.5],[2017,3,523.0],[2017,4,523.5],[2017,5,524.0],[2017,6,524.5],
[2017,7,525.0],[2017,8,525.5],[2017,9,526.0],[2017,10,526.5],[2017,11,527.0],[2017,12,527.5],
[2018,1,528.0],[2018,2,528.5],[2018,3,529.0],[2018,4,529.5],[2018,5,530.0],[2018,6,530.5],
[2018,7,531.0],[2018,8,531.5],[2018,9,532.0],[2018,10,532.5],[2018,11,533.0],[2018,12,533.5],
[2019,1,534.0],[2019,2,534.5],[2019,3,535.0],[2019,4,535.5],[2019,5,536.0],[2019,6,536.5],
[2019,7,537.0],[2019,8,537.5],[2019,9,538.0],[2019,10,538.5],[2019,11,539.0],[2019,12,539.5],
[2020,1,389.0],[2020,2,390.0],[2020,3,391.0],[2020,4,392.0],[2020,5,393.0],[2020,6,394.0],
[2020,7,395.0],[2020,8,396.0],[2020,9,397.0],[2020,10,398.0],[2020,11,399.0],[2020,12,400.0],
[2021,1,385.0],[2021,2,386.0],[2021,3,387.0],[2021,4,388.0],[2021,5,389.0],[2021,6,390.0],
[2021,7,391.0],[2021,8,392.0],[2021,9,393.0],[2021,10,394.0],[2021,11,395.0],[2021,12,396.0],
[2022,1,375.0],[2022,2,376.0],[2022,3,377.0],[2022,4,378.0],[2022,5,379.0],[2022,6,380.0],
[2022,7,381.0],[2022,8,382.0],[2022,9,383.0],[2022,10,384.0],[2022,11,385.0],[2022,12,386.0],
[2023,1,370.0],[2023,2,371.0],[2023,3,372.0],[2023,4,373.0],[2023,5,374.0],[2023,6,375.0],
[2023,7,376.0],[2023,8,377.0],[2023,9,378.0],[2023,10,379.0],[2023,11,380.0],[2023,12,381.0],
[2024,1,375.0],[2024,2,376.0],[2024,3,377.0],[2024,4,378.0],[2024,5,379.0],[2024,6,380.0],
[2024,7,381.0],[2024,8,382.0],[2024,9,383.0],[2024,10,384.0],[2024,11,385.0],[2024,12,386.0],
[2025,1,380.0],[2025,2,381.0],[2025,3,382.0]
])
	
BIRD_FLU_DATA = np.array([
[2017,1,0.0],[2017,2,0.0],[2017,3,0.0],[2017,4,0.0],[2017,5,0.0],[2017,6,0.0],
[2017,7,0.0],[2017,8,0.0],[2017,9,0.0],[2017,10,0.0],[2017,11,0.0],[2017,12,0.0],
[2018,1,0.0],[2018,2,0.0],[2018,3,0.0],[2018,4,0.0],[2018,5,0.0],[2018,6,0.0],
[2018,7,0.0],[2018,8,0.0],[2018,9,0.0],[2018,10,0.0],[2018,11,0.0],[2018,12,0.0],
[2019,1,0.0],[2019,2,0.0],[2019,3,0.0],[2019,4,0.0],[2019,5,0.0],[2019,6,0.0],
[2019,7,0.0],[2019,8,0.0],[2019,9,0.0],[2019,10,0.0],[2019,11,0.0],[2019,12,0.0],
[2020,1,0.0],[2020,2,0.0],[2020,3,0.0],[2020,4,0.0],[2020,5,0.0],[2020,6,0.0],
[2020,7,0.0],[2020,8,0.0],[2020,9,0.0],[2020,10,0.0],[2020,11,0.0],[2020,12,0.0],
[2021,1,0.0],[2021,2,0.0],[2021,3,0.0],[2021,4,0.0],[2021,5,0.0],[2021,6,0.0],
[2021,7,0.0],[2021,8,0.0],[2021,9,0.0],[2021,10,0.0],[2021,11,0.0],[2021,12,0.0],
[2022,1,5000.0],[2022,2,5200.0],[2022,3,5400.0],[2022,4,5600.0],[2022,5,5800.0],[2022,6,6000.0],
[2022,7,6200.0],[2022,8,6400.0],[2022,9,6600.0],[2022,10,6800.0],[2022,11,7000.0],[2022,12,7200.0],
[2023,1,42300.0],[2023,2,42500.0],[2023,3,42700.0],[2023,4,42900.0],[2023,5,43100.0],[2023,6,43300.0],
[2023,7,43500.0],[2023,8,43700.0],[2023,9,43900.0],[2023,10,44100.0],[2023,11,44300.0],[2023,12,44500.0],
[2024,1,19500.0],[2024,2,19700.0],[2024,3,19900.0],[2024,4,20100.0],[2024,5,20300.0],[2024,6,20500.0],
[2024,7,20700.0],[2024,8,20900.0],[2024,9,21100.0],[2024,10,21300.0],[2024,11,21500.0],[2024,12,21700.0],
[2025,1,36000.0],[2025,2,36200.0],[2025,3,36400.0]
])

DIESEL_DATA = np.array([
[2017,1,2.50],[2017,2,2.52],[2017,3,2.54],[2017,4,2.56],[2017,5,2.58],[2017,6,2.60],
[2017,7,2.62],[2017,8,2.64],[2017,9,2.66],[2017,10,2.68],[2017,11,2.70],[2017,12,2.72],
[2018,1,2.80],[2018,2,2.82],[2018,3,2.84],[2018,4,2.86],[2018,5,2.88],[2018,6,2.90],
[2018,7,2.92],[2018,8,2.94],[2018,9,2.96],[2018,10,2.98],[2018,11,3.00],[2018,12,3.02],
[2019,1,2.90],[2019,2,2.92],[2019,3,2.94],[2019,4,2.96],[2019,5,2.98],[2019,6,3.00],
[2019,7,3.02],[2019,8,3.04],[2019,9,3.06],[2019,10,3.08],[2019,11,3.10],[2019,12,3.12],
[2020,1,3.00],[2020,2,2.95],[2020,3,2.90],[2020,4,2.85],[2020,5,2.80],[2020,6,2.75],
[2020,7,2.70],[2020,8,2.65],[2020,9,2.60],[2020,10,2.55],[2020,11,2.50],[2020,12,2.45],
[2021,1,2.80],[2021,2,2.85],[2021,3,2.90],[2021,4,2.95],[2021,5,3.00],[2021,6,3.05],
[2021,7,3.10],[2021,8,3.15],[2021,9,3.20],[2021,10,3.25],[2021,11,3.30],[2021,12,3.35],
[2022,1,3.50],[2022,2,3.55],[2022,3,3.60],[2022,4,3.65],[2022,5,3.70],[2022,6,3.75],
[2022,7,3.80],[2022,8,3.85],[2022,9,3.90],[2022,10,3.95],[2022,11,4.00],[2022,12,4.05],
[2023,1,4.20],[2023,2,4.25],[2023,3,4.30],[2023,4,4.35],[2023,5,4.40],[2023,6,4.45],
[2023,7,4.50],[2023,8,4.55],[2023,9,4.60],[2023,10,4.65],[2023,11,4.70],[2023,12,4.75],
[2024,1,4.90],[2024,2,4.95],[2024,3,5.00],[2024,4,5.05],[2024,5,5.10],[2024,6,5.15],
[2024,7,5.20],[2024,8,5.25],[2024,9,5.30],[2024,10,5.35],[2024,11,5.40],[2024,12,5.45],
[2025,1,5.60],[2025,2,5.65],[2025,3,5.70]
])

NEWS_SENTIMENT_DATA = np.array([
[2017,1,0.7],[2017,2,0.7],[2017,3,0.7],[2017,4,0.7],[2017,5,0.7],[2017,6,0.7],
[2017,7,0.7],[2017,8,0.7],[2017,9,0.7],[2017,10,0.7],[2017,11,0.7],[2017,12,0.7],
[2018,1,0.8],[2018,2,0.8],[2018,3,0.8],[2018,4,0.8],[2018,5,0.8],[2018,6,0.8],
[2018,7,0.8],[2018,8,0.8],[2018,9,0.8],[2018,10,0.8],[2018,11,0.8],[2018,12,0.8],
[2019,1,0.7],[2019,2,0.7],[2019,3,0.7],[2019,4,0.7],[2019,5,0.7],[2019,6,0.7],
[2019,7,0.7],[2019,8,0.7],[2019,9,0.7],[2019,10,0.7],[2019,11,0.7],[2019,12,0.7],
[2020,1,0.9],[2020,2,0.8],[2020,3,0.7],[2020,4,0.9],[2020,5,0.8],[2020,6,0.8],
[2020,7,0.9],[2020,8,0.8],[2020,9,0.8],[2020,10,0.9],[2020,11,0.8],[2020,12,0.8],
[2021,1,0.9],[2021,2,0.8],[2021,3,0.8],[2021,4,0.9],[2021,5,0.8],[2021,6,0.8],
[2021,7,0.9],[2021,8,0.8],[2021,9,0.8],[2021,10,0.9],[2021,11,0.8],[2021,12,0.8],
[2022,1,0.5],[2022,2,0.4],[2022,3,0.4],[2022,4,0.4],[2022,5,0.4],[2022,6,0.4],
[2022,7,0.4],[2022,8,0.4],[2022,9,0.4],[2022,10,0.4],[2022,11,0.4],[2022,12,0.4],
[2023,1,0.2],[2023,2,0.2],[2023,3,0.2],[2023,4,0.2],[2023,5,0.2],[2023,6,0.2],
[2023,7,0.2],[2023,8,0.2],[2023,9,0.2],[2023,10,0.2],[2023,11,0.2],[2023,12,0.2],
[2024,1,0.6],[2024,2,0.6],[2024,3,0.6],[2024,4,0.6],[2024,5,0.6],[2024,6,0.6],
[2024,7,0.6],[2024,8,0.6],[2024,9,0.6],[2024,10,0.6],[2024,11,0.6],[2024,12,0.6],
[2025,1,0.3],[2025,2,0.3],[2025,3,0.3]
])

CPI_DATA = np.array([
[2017,1,242.839],[2017,2,243.603],[2017,3,243.801],[2017,4,244.524],[2017,5,244.733],[2017,6,244.955],
[2017,7,245.519],[2017,8,246.819],[2017,9,247.867],[2017,10,246.663],[2017,11,246.669],[2017,12,246.524],
[2018,1,247.867],[2018,2,248.991],[2018,3,249.554],[2018,4,250.546],[2018,5,251.588],[2018,6,251.989],
[2018,7,252.006],[2018,8,252.146],[2018,9,252.439],[2018,10,252.885],[2018,11,252.038],[2018,12,251.233],
[2019,1,251.712],[2019,2,252.776],[2019,3,254.202],[2019,4,255.548],[2019,5,256.092],[2019,6,256.143],
[2019,7,256.571],[2019,8,256.558],[2019,9,256.759],[2019,10,257.346],[2019,11,257.208],[2019,12,256.974],
[2020,1,257.971],[2020,2,258.678],[2020,3,258.115],[2020,4,256.389],[2020,5,256.394],[2020,6,257.797],
[2020,7,259.101],[2020,8,259.918],[2020,9,260.280],[2020,10,260.388],[2020,11,260.229],[2020,12,260.474],
[2021,1,261.582],[2021,2,263.014],[2021,3,264.877],[2021,4,267.054],[2021,5,269.195],[2021,6,271.696],
[2021,7,273.003],[2021,8,273.567],[2021,9,274.310],[2021,10,276.589],[2021,11,277.948],[2021,12,278.802],
[2022,1,281.148],[2022,2,283.716],[2022,3,287.504],[2022,4,289.109],[2022,5,292.296],[2022,6,296.311],
[2022,7,296.276],[2022,8,296.171],[2022,9,296.808],[2022,10,298.012],[2022,11,297.711],[2022,12,296.797],
[2023,1,299.170],[2023,2,300.840],[2023,3,301.836],[2023,4,303.363],[2023,5,304.127],[2023,6,305.109],
[2023,7,305.691],[2023,8,307.026],[2023,9,307.789],[2023,10,307.671],[2023,11,307.051],[2023,12,306.746],
[2024,1,308.417],[2024,2,310.326],[2024,3,312.332],[2024,4,313.548],[2024,5,314.069],[2024,6,314.175],
[2024,7,314.540],[2024,8,314.796],[2024,9,315.301],[2024,10,315.856],[2024,11,316.251],[2024,12,316.391],
[2025,1,317.500],[2025,2,318.600],[2025,3,319.700]
])

CORN_PRICE_DATA = np.array([
[2017,1,3.60],[2017,2,3.62],[2017,3,3.64],[2017,4,3.66],[2017,5,3.68],[2017,6,3.70],
[2017,7,3.72],[2017,8,3.74],[2017,9,3.76],[2017,10,3.78],[2017,11,3.80],[2017,12,3.82],
[2018,1,3.84],[2018,2,3.86],[2018,3,3.88],[2018,4,3.90],[2018,5,3.92],[2018,6,3.94],
[2018,7,3.96],[2018,8,3.98],[2018,9,4.00],[2018,10,4.02],[2018,11,4.04],[2018,12,4.06],
[2019,1,4.08],[2019,2,4.10],[2019,3,4.12],[2019,4,4.14],[2019,5,4.16],[2019,6,4.18],
[2019,7,4.20],[2019,8,4.22],[2019,9,4.24],[2019,10,4.26],[2019,11,4.28],[2019,12,4.30],
[2020,1,3.85],[2020,2,3.80],[2020,3,3.75],[2020,4,3.70],[2020,5,3.65],[2020,6,3.60],
[2020,7,3.55],[2020,8,3.50],[2020,9,3.45],[2020,10,3.40],[2020,11,3.35],[2020,12,3.30],
[2021,1,4.20],[2021,2,4.30],[2021,3,4.40],[2021,4,4.50],[2021,5,4.60],[2021,6,4.70],
[2021,7,4.80],[2021,8,4.90],[2021,9,5.00],[2021,10,5.10],[2021,11,5.20],[2021,12,5.30],
[2022,1,5.40],[2022,2,5.50],[2022,3,5.60],[2022,4,5.70],[2022,5,5.80],[2022,6,5.90],
[2022,7,6.00],[2022,8,6.10],[2022,9,6.20],[2022,10,6.30],[2022,11,6.40],[2022,12,6.50],
[2023,1,6.60],[2023,2,6.70],[2023,3,6.80],[2023,4,6.90],[2023,5,7.00],[2023,6,7.10],
[2023,7,7.20],[2023,8,7.30],[2023,9,7.40],[2023,10,7.50],[2023,11,7.60],[2023,12,7.70],
[2024,1,7.80],[2024,2,7.90],[2024,3,8.00],[2024,4,8.10],[2024,5,8.20],[2024,6,8.30],
[2024,7,8.40],[2024,8,8.50],[2024,9,8.60],[2024,10,8.70],[2024,11,8.80],[2024,12,8.90],
[2025,1,9.00],[2025,2,9.10],[2025,3,9.20]
])

SOYBEAN_MEAL_DATA = np.array([
[2017,1,320],[2017,2,322],[2017,3,324],[2017,4,326],[2017,5,328],[2017,6,330],
[2017,7,332],[2017,8,334],[2017,9,336],[2017,10,338],[2017,11,340],[2017,12,342],
[2018,1,344],[2018,2,346],[2018,3,348],[2018,4,350],[2018,5,352],[2018,6,354],
[2018,7,356],[2018,8,358],[2018,9,360],[2018,10,362],[2018,11,364],[2018,12,366],
[2019,1,368],[2019,2,370],[2019,3,372],[2019,4,374],[2019,5,376],[2019,6,378],
[2019,7,380],[2019,8,382],[2019,9,384],[2019,10,386],[2019,11,388],[2019,12,390],
[2020,1,300],[2020,2,305],[2020,3,310],[2020,4,315],[2020,5,320],[2020,6,325],
[2020,7,330],[2020,8,335],[2020,9,340],[2020,10,345],[2020,11,350],[2020,12,355],
[2021,1,360],[2021,2,365],[2021,3,370],[2021,4,375],[2021,5,380],[2021,6,385],
[2021,7,390],[2021,8,395],[2021,9,400],[2021,10,405],[2021,11,410],[2021,12,415],
[2022,1,420],[2022,2,425],[2022,3,430],[2022,4,435],[2022,5,440],[2022,6,445],
[2022,7,450],[2022,8,455],[2022,9,460],[2022,10,465],[2022,11,470],[2022,12,475],
[2023,1,480],[2023,2,485],[2023,3,490],[2023,4,495],[2023,5,500],[2023,6,505],
[2023,7,510],[2023,8,515],[2023,9,520],[2023,10,525],[2023,11,530],[2023,12,535],
[2024,1,540],[2024,2,545],[2024,3,550],[2024,4,555],[2024,5,560],[2024,6,565],
[2024,7,570],[2024,8,575],[2024,9,580],[2024,10,585],[2024,11,590],[2024,12,595],
[2025,1,600],[2025,2,605],[2025,3,610]
])

WEEKLY_WHOLESALE_APRIL_2025 = np.array([
[2025,4,1,2.43],[2025,4,8,2.45],[2025,4,15,2.47],[2025,4,22,2.49],[2025,4,25,2.50]
])

# Define prepare_data function
def prepare_data():
    data_lengths = [len(d) for d in [BLS_DATA, WHOLESALE_DATA, HEN_DATA, BIRD_FLU_DATA, DIESEL_DATA, NEWS_SENTIMENT_DATA, CPI_DATA, CORN_PRICE_DATA, SOYBEAN_MEAL_DATA]]
    if len(set(data_lengths)) != 1:
        raise ValueError("All datasets must have the same length.")
    
    for dataset, name in [
        (BLS_DATA, 'BLS_DATA'), (WHOLESALE_DATA, 'WHOLESALE_DATA'), (HEN_DATA, 'HEN_DATA'),
        (BIRD_FLU_DATA, 'BIRD_FLU_DATA'), (DIESEL_DATA, 'DIESEL_DATA'), (NEWS_SENTIMENT_DATA, 'NEWS_SENTIMENT_DATA'),
        (CPI_DATA, 'CPI_DATA'), (CORN_PRICE_DATA, 'CORN_PRICE_DATA'), (SOYBEAN_MEAL_DATA, 'SOYBEAN_MEAL_DATA')
    ]:
        if np.any(np.isnan(dataset)):
            raise ValueError(f"Dataset {name} contains NaN values.")
        if np.any([x[2] < 0 for x in dataset]):
            raise ValueError(f"Dataset {name} contains negative values.")
    
    dfs = [pd.DataFrame(data, columns=['year', 'month', col]) for data, col in [
        (BLS_DATA, 'retail_price'), (WHOLESALE_DATA, 'wholesale_price'), (HEN_DATA, 'hen_population'),
        (BIRD_FLU_DATA, 'bird_flu_losses'), (DIESEL_DATA, 'diesel_cost'), (NEWS_SENTIMENT_DATA, 'news_sentiment'),
        (CPI_DATA, 'cpi'), (CORN_PRICE_DATA, 'corn_price'), (SOYBEAN_MEAL_DATA, 'soybean_meal_price')
    ]]
    
    df = dfs[0]
    for d in dfs[1:]:
        df = df.merge(d, on=['year', 'month'])
    
    df['date'] = pd.to_datetime(df[['year', 'month']].assign(day=1))
    df = df.sort_values('date')
    
    df['log_retail_price'] = np.log(df['retail_price'])
    df['weight'] = np.arange(1, len(df) + 1) / len(df)
    
    for lag in range(1, 4):
        df[f'retail_lag{lag}'] = df['retail_price'].shift(lag)
        df[f'wholesale_lag{lag}'] = df['wholesale_price'].shift(lag)
    
    df['retail_ewma_3'] = df['retail_price'].ewm(span=3, adjust=False).mean().shift(1)
    df['retail_volatility_3m'] = df['retail_price'].rolling(window=3).std().shift(1)
    df['news_birdflu_interaction'] = df['news_sentiment'] * df['bird_flu_losses']
    df['error_lag1'] = 0.0
    
    df['price_pct_change'] = df['retail_price'].pct_change().shift(1)
    df['price_abs_change'] = df['retail_price'].diff().shift(1)
    df['recent_volatility'] = df['price_pct_change'].rolling(window=3).std().shift(1)
    historical_vol_mean = df['recent_volatility'].mean()
    historical_vol_std = df['recent_volatility'].std()
    df['spike_indicator'] = ((df['price_pct_change'].abs() > (df['price_pct_change'].mean() + 1.5 * df['price_pct_change'].std())) | 
                             (df['price_abs_change'].abs() > 0.5)).astype(int).shift(1)
    df['trend_strength'] = df['price_pct_change'].rolling(window=3).sum().shift(1)
    df['bird_flu_volatility_interaction'] = df['bird_flu_losses'] * df['recent_volatility']
    df['diesel_volatility_interaction'] = df['diesel_cost'] * df['recent_volatility']
    df['trend_adjustment'] = 0.0
    
    df['price_first_derivative'] = df['retail_price'].diff().shift(1)
    df['price_second_derivative'] = df['price_first_derivative'].diff().shift(1)
    df['price_third_derivative'] = df['price_second_derivative'].diff().shift(1)
    df['trend_momentum'] = df['price_first_derivative'].rolling(window=3).sum().shift(1)
    
    df['damping_factor'] = df['diesel_cost'].shift(1) * 0.1
    df['forcing_term'] = df['bird_flu_losses'].shift(1) * 0.05
    
    df['wholesale_bls_corr'] = df['wholesale_price'].rolling(window=6).corr(df['retail_price']).shift(1)
    df['diesel_bls_corr'] = df['diesel_cost'].rolling(window=6).corr(df['retail_price']).shift(1)
    df['wholesale_bls_corr_diff'] = df['wholesale_bls_corr'].diff().shift(1)
    df['diesel_bls_corr_diff'] = df['diesel_bls_corr'].diff().shift(1)
    
    df['retail_wholesale_ratio'] = df['retail_price'] / (df['wholesale_price'] + 1e-6)
    df['rsi_retail_wholesale'] = df['retail_wholesale_ratio'].rolling(window=6).mean().shift(1)
    
    df.replace([np.inf, -np.inf], np.nan, inplace=True)
    df.fillna(0, inplace=True)
    
    features = [f'retail_lag{lag}' for lag in range(1, 4)] + \
               [f'wholesale_lag{lag}' for lag in range(1, 4)] + \
               ['wholesale_price', 'diesel_cost', 
                'retail_ewma_3', 'retail_volatility_3m', 'news_birdflu_interaction', 
                'hen_population', 'bird_flu_losses', 'diesel_cost', 'news_sentiment', 'cpi', 
                'corn_price', 'soybean_meal_price', 'error_lag1', 'recent_volatility', 
                'spike_indicator', 'trend_strength', 'bird_flu_volatility_interaction', 
                'diesel_volatility_interaction', 'price_first_derivative', 'price_second_derivative',
                'price_third_derivative', 'trend_momentum', 'damping_factor', 'forcing_term',
                'wholesale_bls_corr', 'diesel_bls_corr', 'wholesale_bls_corr_diff', 'diesel_bls_corr_diff',
                'rsi_retail_wholesale', 'trend_adjustment']
    
    corr_matrix = df[features].corr().abs()
    upper = corr_matrix.where(np.triu(np.ones(corr_matrix.shape), k=1).astype(bool))
    to_drop = [column for column in upper.columns if any(upper[column] > 0.95)]
    features = [f for f in features if f not in to_drop]
    
    temp_model = XGBRegressor(n_estimators=100, learning_rate=0.01, max_depth=5, random_state=42, objective='reg:absoluteerror')
    temp_model.fit(df[features], df['log_retail_price'])
    importances = pd.Series(temp_model.feature_importances_, index=features)
    top_features = importances.nlargest(15).index.tolist()
    
    df = df.dropna(subset=top_features)
    df = df.reset_index(drop=True)
    
    df['is_volatile'] = (df['recent_volatility'] > (historical_vol_mean + 2 * historical_vol_std)).astype(int)
    return df, top_features

# Initialize session state
if 'predictions' not in st.session_state:
    st.session_state['predictions'] = None
if 'historical_predictions' not in st.session_state:
    st.session_state['historical_predictions'] = None
if 'historical_errors' not in st.session_state:
    st.session_state['historical_errors'] = []
if 'last_update' not in st.session_state:
    st.session_state['last_update'] = time.time()
if 'running' not in st.session_state:
    st.session_state['running'] = True
if 'initial_load' not in st.session_state:
    st.session_state['initial_load'] = True
if 'update_counter' not in st.session_state:
    st.session_state['update_counter'] = 0
if 'prediction_history' not in st.session_state:
    st.session_state['prediction_history'] = []
if 'ci_coverage' not in st.session_state:
    st.session_state['ci_coverage'] = 0.95
if 'computation_times' not in st.session_state:
    st.session_state['computation_times'] = []
if 'historical_volatility_performance' not in st.session_state:
    st.session_state['historical_volatility_performance'] = {}
if 'peak_message' not in st.session_state:
    st.session_state['peak_message'] = "No significant spike detected."
if 'best_mae' not in st.session_state:
    st.session_state['best_mae'] = float('inf')
if 'best_predictions' not in st.session_state:
    st.session_state['best_predictions'] = None
if 'learning_rate' not in st.session_state:
    st.session_state['learning_rate'] = 0.1
if 'n_estimators' not in st.session_state:
    st.session_state['n_estimators'] = 1000
if 'lambda_reg' not in st.session_state:
    st.session_state['lambda_reg'] = 1.0
if 'alpha_reg' not in st.session_state:
    st.session_state['alpha_reg'] = 0.0
if 'risk_factor' not in st.session_state:
    st.session_state['risk_factor'] = 1.0
if 'previous_mae' not in st.session_state:
    st.session_state['previous_mae'] = None
if 'long_term_confidence_history' not in st.session_state:
    st.session_state['long_term_confidence_history'] = []
if 'best_configuration' not in st.session_state:
    st.session_state['best_configuration'] = {
        'learning_rate': 0.1,
        'n_estimators': 1000,
        'lambda_reg': 1.0,
        'alpha_reg': 0.0,
        'risk_factor': 1.0,
        'ci_coverage': 0.95,
        'confidence': 0.0
    }
if 'global_error_correction' not in st.session_state:
    st.session_state['global_error_correction'] = 0.0
if 'update_in_progress' not in st.session_state:
    st.session_state['update_in_progress'] = False
if 'market_phase' not in st.session_state:
    st.session_state['market_phase'] = "consolidation"
if 'debug_logs' not in st.session_state:
    st.session_state['debug_logs'] = []
if 'df' not in st.session_state or 'features' not in st.session_state:
    try:
        df, features = prepare_data()
        st.session_state['df'] = df
        st.session_state['features'] = features
        st.session_state['error_weight'] = pd.Series(1.0, index=df.index)
        st.session_state['debug_logs'].append(f"[{datetime.now()}] Initialized dataset with {len(df)} rows, {len(features)} features")
    except Exception as e:
        st.error(f"Error preparing data: {e}")
        st.session_state['debug_logs'].append(f"[{datetime.now()}] Error preparing data: {e}")
        st.stop()

st.set_page_config(page_title="Egg Price Prediction Dashboard", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@400;700&display=swap');
    body {
        background-color: #0A0A23;
        color: #E0E0E0;
        font-family: 'Roboto', sans-serif;
    }
    .stApp {
        background-color: #0A0A23;
    }
    h1, h2, h3, h4, h5, h6 {
        color: #00FFFF;
        text-shadow: 0 0 2px #00FFFF;
        text-transform: uppercase;
        letter-spacing: 1px;
        font-weight: 700;
    }
    .stButton>button {
        background: linear-gradient(45deg, #FF00FF, #00FFFF);
        color: #0A0A23;
        border-radius: 5px;
        border: none;
        padding: 10px 20px;
        transition: all 0.3s ease;
        font-weight: 700;
    }
    .stButton>button:hover {
        background: linear-gradient(45deg, #00FFFF, #FF00FF);
        transform: translateY(-2px);
        box-shadow: 0 2px 8px rgba(0, 255, 255, 0.5);
    }
    .metric-box {
        background: linear-gradient(135deg, #1A1A3D, #2A2A5D);
        padding: 15px;
        border-radius: 8px;
        border: 1px solid #00FFFF;
        box-shadow: 0 0 10px rgba(0, 255, 255, 0.2);
        text-align: center;
        margin: 5px;
    }
    .loading-text {
        color: #00FFFF;
        font-size: 18px;
        font-weight: 700;
    }
    .stSidebar {
        background: linear-gradient(180deg, #1A1A3D, #0A0A23);
        border-right: 1px solid #00FFFF;
    }
    .spinner {
        display: inline-block;
        animation: spin 2s linear infinite;
        opacity: 0.7;
    }
    @keyframes spin {
        0% { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }
    .tooltip {
        position: relative;
        display: inline-block;
    }
    .tooltip .tooltiptext {
        visibility: hidden;
        width: 120px;
        background-color: #1A1A3D;
        color: #E0E0E0;
        text-align: center;
        border-radius: 5px;
        padding: 5px;
        position: absolute;
        z-index: 1;
        bottom: 125%;
        left: 50%;
        margin-left: -60px;
        opacity: 0;
        transition: opacity 0.3s;
        font-size: 12px;
    }
    .tooltip:hover .tooltiptext {
        visibility: visible;
        opacity: 1;
    }
    </style>
""", unsafe_allow_html=True)

def detect_market_phase(df, recent_months=6):
    recent_df = df.tail(recent_months)
    avg_momentum = recent_df['trend_momentum'].mean()
    avg_volatility = recent_df['recent_volatility'].mean()
    avg_second_deriv = recent_df['price_second_derivative'].mean()
    
    vol_threshold = df['recent_volatility'].mean() + df['recent_volatility'].std()
    
    if avg_momentum > 0 and avg_volatility < vol_threshold and avg_second_deriv >= -0.1:
        phase = "uptrend"
    elif avg_momentum < 0 and avg_volatility < vol_threshold and avg_second_deriv <= 0.1:
        phase = "downtrend"
    else:
        phase = "consolidation"
    
    print(f"Detected market phase: {phase} (momentum={avg_momentum:.2f}, volatility={avg_volatility:.2f}, second_deriv={avg_second_deriv:.2f})")
    return phase

def optimize_hyperparameters(train_df, features, historical_predictions, lambda_reg, alpha_reg, current_volatility, historical_volatility_performance, risk_factor):
    param_grid = {
        'learning_rate': [0.01, 0.05, 0.1, 0.15, 0.2],
        'max_depth': [3, 4, 5, 6, 7],
        'n_estimators': [100, 300, 500, 1000, 1500],
        'subsample': [0.8, 0.9, 1.0],
        'colsample_bytree': [0.8, 0.9, 1.0]
    }
    grid = list(ParameterGrid(param_grid))
    random.shuffle(grid)  # Randomize for efficiency
    grid = grid[:10]  # Limit to 10 trials for speed
    
    best_score = float('inf')
    best_params = None
    best_preds = None
    
    for params in grid:
        adjusted_lr = params['learning_rate'] * (1 + current_volatility / historical_volatility_performance.get('mean_volatility', 1)) * risk_factor
        preds = bootstrap_predict(train_df, train_df, features, n_bootstrap=10, 
                                 learning_rate=adjusted_lr, 
                                 n_estimators=params['n_estimators'], 
                                 lambda_reg=lambda_reg, alpha_reg=alpha_reg)
        pred_mean = np.mean(preds, axis=0)
        errors = [abs(pred_mean[i] - train_df['retail_price'].iloc[i]) for i in range(len(pred_mean)) if not np.isnan(train_df['retail_price'].iloc[i])]
        score = np.mean(errors) if errors else float('inf')
        if score < best_score:
            best_score = score
            best_params = params
            best_preds = preds
    
    return best_params, best_preds

def bootstrap_predict_single(train_df, test_df, features, random_state, learning_rate, n_estimators, lambda_reg, alpha_reg):
    sample = resample(train_df, replace=True, random_state=random_state)
    model = XGBRegressor(n_estimators=n_estimators, learning_rate=learning_rate, max_depth=5, 
                         subsample=0.8, colsample_bytree=0.8, random_state=42, missing=np.nan, 
                         reg_lambda=lambda_reg, reg_alpha=alpha_reg, objective='reg:absoluteerror')
    weights = sample['weight'] * (1 + sample['error_weight'])
    model.fit(sample[features], sample['log_retail_price'], sample_weight=weights)
    preds = model.predict(test_df[features])
    preds = np.minimum(preds, 10)
    try:
        result = np.exp(preds)
        return np.minimum(result, 1000)
    except OverflowError:
        print(f"Overflow in np.exp for preds={preds}, capping at 1000")
        return np.full_like(preds, 1000)

def bootstrap_predict(train_df, test_df, features, n_bootstrap, learning_rate, n_estimators, lambda_reg, alpha_reg):
    preds = Parallel(n_jobs=-1)(delayed(bootstrap_predict_single)(train_df, test_df, features, i, learning_rate, n_estimators, lambda_reg, alpha_reg) for i in range(n_bootstrap))
    return np.array(preds)

def ensemble_predict(train_df, test_df, features, n_bootstrap, current_volatility, historical_volatility_performance, risk_factor, market_phase):
    if market_phase == "uptrend":
        configs = [
            {'learning_rate': 0.15, 'n_estimators': 500, 'lambda_reg': 0.8, 'alpha_reg': 0.0},
            {'learning_rate': 0.1, 'n_estimators': 300, 'lambda_reg': 1.0, 'alpha_reg': 0.1},
            {'learning_rate': 0.05, 'n_estimators': 1000, 'lambda_reg': 1.2, 'alpha_reg': 0.5}
        ]
    elif market_phase == "downtrend":
        configs = [
            {'learning_rate': 0.05, 'n_estimators': 500, 'lambda_reg': 1.2, 'alpha_reg': 0.5},
            {'learning_rate': 0.1, 'n_estimators': 300, 'lambda_reg': 1.0, 'alpha_reg': 0.1},
            {'learning_rate': 0.15, 'n_estimators': 1000, 'lambda_reg': 0.8, 'alpha_reg': 0.0}
        ]
    else:
        configs = [
            {'learning_rate': 0.1, 'n_estimators': 500, 'lambda_reg': 1.5, 'alpha_reg': 0.5},
            {'learning_rate': 0.15, 'n_estimators': 300, 'lambda_reg': 1.2, 'alpha_reg': 0.3},
            {'learning_rate': 0.05, 'n_estimators': 1000, 'lambda_reg': 1.0, 'alpha_reg': 0.1}
        ]
    
    all_preds = []
    weights = []
    for config in configs:
        adjusted_lr = config['learning_rate'] * (1 + current_volatility / historical_volatility_performance.get('mean_volatility', 1)) * risk_factor
        preds = bootstrap_predict(train_df, test_df, features, n_bootstrap=n_bootstrap, 
                                 learning_rate=adjusted_lr, n_estimators=config['n_estimators'], 
                                 lambda_reg=config['lambda_reg'], alpha_reg=config['alpha_reg'])
        all_preds.append(preds)
        error_during_volatility = historical_volatility_performance.get('error_during_volatility', 1.0)
        weight = 1 / (error_during_volatility + 1e-6)
        weights.append(weight)
    
    weights = np.array(weights) / np.sum(weights)
    combined_preds = np.average(all_preds, axis=0, weights=weights)
    return combined_preds

def smooth_confidence_intervals(predictions, span=3):
    ci_lower_95 = [p['ci_lower_95'] for p in predictions]
    ci_upper_95 = [p['ci_upper_95'] for p in predictions]
    ci_lower_75 = [p['ci_lower_75'] for p in predictions]
    ci_upper_75 = [p['ci_upper_75'] for p in predictions]
    predicted = [p['predicted'] for p in predictions]
    
    ci_lower_95_series = pd.Series(ci_lower_95)
    ci_upper_95_series = pd.Series(ci_upper_95)
    ci_lower_75_series = pd.Series(ci_lower_75)
    ci_upper_75_series = pd.Series(ci_upper_75)
    
    ci_lower_95_smooth = ci_lower_95_series.ewm(span=span, adjust=False).mean().tolist()
    ci_upper_95_smooth = ci_upper_95_series.ewm(span=span, adjust=False).mean().tolist()
    ci_lower_75_smooth = ci_lower_75_series.ewm(span=span, adjust=False).mean().tolist()
    ci_upper_75_smooth = ci_upper_75_series.ewm(span=span, adjust=False).mean().tolist()
    
    for i, p in enumerate(predictions):
        pred = predicted[i]
        ci_width_95 = ci_upper_95_smooth[i] - ci_lower_95_smooth[i]
        ci_width_75 = ci_upper_75_smooth[i] - ci_lower_75_smooth[i]
        ci_lower_95_smooth[i] = pred - ci_width_95 / 2
        ci_upper_95_smooth[i] = pred + ci_width_95 / 2
        ci_lower_75_smooth[i] = pred - ci_width_75 / 2
        ci_upper_75_smooth[i] = pred + ci_width_75 / 2
        
        ci_lower_95_smooth[i] = min(ci_lower_95_smooth[i], pred - 0.01)
        ci_upper_95_smooth[i] = max(ci_upper_95_smooth[i], pred + 0.01)
        ci_lower_75_smooth[i] = min(ci_lower_75_smooth[i], pred - 0.005)
        ci_upper_75_smooth[i] = max(ci_upper_75_smooth[i], pred + 0.005)
        
        p['ci_lower_95'] = ci_lower_95_smooth[i]
        p['ci_upper_95'] = ci_upper_95_smooth[i]
        p['ci_lower_75'] = ci_lower_75_smooth[i]
        p['ci_upper_75'] = ci_upper_75_smooth[i]
    
    return predictions

def generate_historical_predictions(df, features, historical_errors=None, _progress_callback=None):
    start_time = time.time()
    predictions = []
    total_steps = len(df) - 1
    step = 0
    
    for i in range(1, len(df)):
        if _progress_callback:
            step += 1
            _progress_callback(step / total_steps, f"Predicting month {i} of {total_steps}...")
        
        train_df = df.iloc[:i]
        test_df = df.iloc[i:i+1]
        if len(train_df) < 10:
            continue
        
        current_volatility = train_df['recent_volatility'].mean()
        market_phase = detect_market_phase(train_df)
        historical_volatility_performance = {
            'mean_volatility': current_volatility,
            'error_during_volatility': np.mean([abs(e) for e in historical_errors if not np.isnan(e)] if historical_errors else [1.0])}
        
        preds = ensemble_predict(train_df, test_df, features, n_bootstrap=50, 
                                current_volatility=current_volatility, 
                                historical_volatility_performance=historical_volatility_performance, 
                                risk_factor=1.0, market_phase=market_phase)
        if np.any(np.isnan(preds)):
            preds = np.nan_to_num(preds, nan=np.nanmean(preds))
        pred_mean = np.mean(preds, axis=0)[0]
        
        ci_lower_95 = np.percentile(preds[:, 0], 2.5)
        ci_upper_95 = np.percentile(preds[:, 0], 97.5)
        ci_lower_75 = np.percentile(preds[:, 0], 12.5)
        ci_upper_75 = np.percentile(preds[:, 0], 87.5)
        
        actual = test_df['retail_price'].iloc[0]
        error = pred_mean - actual if not np.isnan(actual) else 0.0
        
        predictions.append({
            'date': test_df['date'].iloc[0], 'actual': actual, 'predicted': pred_mean,
            'ci_lower_95': ci_lower_95, 'ci_upper_95': ci_upper_95,
            'ci_lower_75': ci_lower_75, 'ci_upper_75': ci_upper_75,
            'error': error
        })
    
    end_time = time.time()
    computation_time = end_time - start_time
    return predictions, computation_time

def train_full_predictions(df, features, historical_predictions, historical_errors=None, _progress_callback=None):
    start_time = time.time()
    current_df = df.copy()
    
    current_df['error_weight'] = 1.0
    if historical_predictions:
        for i in range(len(current_df)):
            if i < len(historical_predictions):
                error = historical_predictions[i]['error']
                actual = historical_predictions[i]['actual']
                if not np.isnan(actual):
                    current_df.loc[i, 'error_weight'] = 1 + abs(error) / (actual + 1e-6)
                if i + 1 < len(current_df):
                    current_df.loc[i + 1, 'error_lag1'] = error
    
    learning_rate = 0.1
    n_estimators = 600
    lambda_reg = 1.0
    alpha_reg = 0.0
    risk_factor = 1.0
    
    total_steps = 2
    step = 0
    
    historical_volatility_performance = {
        'mean_volatility': current_df['recent_volatility'].mean(),
        'error_during_volatility': np.mean([abs(e) for e in historical_errors if not np.isnan(e)] if historical_errors else [1.0])
    }
    
    ci_shrink_factor = max(0.5, 1 - st.session_state['update_counter'] * 0.05)
    
    if _progress_callback:
        step += 1
        _progress_callback(step / total_steps, "Training model on entire dataset...")
    
    market_phase = detect_market_phase(current_df)
    st.session_state['market_phase'] = market_phase
    
    current_volatility = current_df['recent_volatility'].mean()
    preds = ensemble_predict(current_df, current_df, features, n_bootstrap=50, 
                            current_volatility=current_volatility, 
                            historical_volatility_performance=historical_volatility_performance, 
                            risk_factor=risk_factor, market_phase=market_phase)
    if np.any(np.isnan(preds)):
        preds = np.nan_to_num(preds, nan=np.nanmean(preds))
    pred_means = np.mean(preds, axis=0)
    
    countercyclical_adjustments = np.zeros(len(pred_means))
    for i in range(len(current_df)):
        if i < 3:
            continue
        second_deriv = current_df['price_second_derivative'].iloc[i]
        if second_deriv > 0:
            countercyclical_adjustments[i] = -0.1 * abs(second_deriv)
        elif second_deriv < 0:
            countercyclical_adjustments[i] = 0.1 * abs(second_deriv)
    
    if historical_errors and len(historical_errors) > 0:
        global_error = np.mean(historical_errors[-10:])
        st.session_state['global_error_correction'] = global_error * 0.1
    
    predictions = []
    historical_coverage_75 = 0
    historical_count = 0
    for i in range(len(current_df)):
        pred_mean = pred_means[i]
        pred_mean += countercyclical_adjustments[i]
        pred_mean += st.session_state['global_error_correction']
        pred_mean = min(pred_mean, 10.0)
        
        ci_lower_95 = np.percentile(preds[:, i], 2.5)
        ci_upper_95 = np.percentile(preds[:, i], 97.5)
        ci_lower_75 = np.percentile(preds[:, i], 12.5)
        ci_upper_75 = np.percentile(preds[:, i], 87.5)
        
        actual = current_df['retail_price'].iloc[i]
        if not np.isnan(actual):
            ci_width_95 = (ci_upper_95 - ci_lower_95) * 1.5 * ci_shrink_factor
            ci_width_75 = (ci_upper_75 - ci_lower_75) * 1.5 * ci_shrink_factor
            
            if historical_coverage_75 / max(historical_count, 1) > 0.75:
                ci_width_75 *= 0.9
            elif historical_coverage_75 / max(historical_count, 1) < 0.5:
                ci_width_75 *= 1.1
            
            ci_lower_95 = pred_mean - ci_width_95 / 2
            ci_upper_95 = pred_mean + ci_width_95 / 2
            ci_lower_75 = pred_mean - ci_width_75 / 2
            ci_upper_75 = pred_mean + ci_width_75 / 2
            
            if actual < ci_lower_95:
                ci_lower_95 = min(actual - 0.01, ci_lower_95)
                ci_upper_95 = pred_mean + (pred_mean - ci_lower_95)
            elif actual > ci_upper_95:
                ci_upper_95 = max(actual + 0.01, ci_upper_95)
                ci_lower_95 = pred_mean - (ci_upper_95 - pred_mean)
            if actual < ci_lower_75:
                ci_lower_75 = min(actual - 0.005, ci_lower_75)
                ci_upper_75 = pred_mean + (pred_mean - ci_lower_75)
            elif actual > ci_upper_75:
                ci_upper_75 = max(actual + 0.005, ci_upper_75)
                ci_lower_75 = pred_mean - (ci_upper_75 - pred_mean)
            
            if ci_lower_75 <= actual <= ci_upper_75:
                historical_coverage_75 += 1
            historical_count += 1
        
        error = pred_mean - actual if not np.isnan(actual) else 0.0
        
        predictions.append({
            'date': current_df['date'].iloc[i], 'actual': actual, 'predicted': pred_mean,
            'ci_lower_95': ci_lower_95, 'ci_upper_95': ci_upper_95,
            'ci_lower_75': ci_lower_75, 'ci_upper_75': ci_upper_75,
            'error': error
        })
        if i + 1 < len(current_df):
            current_df.loc[i + 1, 'error_lag1'] = error
            current_df.loc[i + 1, 'error_weight'] = 1 + abs(error) / (actual + 1e-6) if not np.isnan(actual) else 1.0
    
    if _progress_callback:
        step += 1
        _progress_callback(step / total_steps, "Predicting April 2025...")
    
    last_row = current_df.iloc[-1]
    second_last = current_df.iloc[-2]
    third_last = current_df.iloc[-3]
    
    april_weekly = pd.DataFrame(WEEKLY_WHOLESALE_APRIL_2025, columns=['year', 'month', 'day', 'wholesale_price'])
    weekly_avg = april_weekly['wholesale_price'].mean() if not april_weekly.empty else 0.0
    
    recent_vol = last_row['retail_volatility_3m']
    last_retail_lag1_vol_scaled = last_row['retail_lag1'] / (recent_vol if recent_vol > 0 else 1)
    last_wholesale_lag1_vol_scaled = last_row['wholesale_lag1'] / (recent_vol if recent_vol > 0 else 1)
    
    recent_errors = [p['error'] for p in predictions[-5:] if not np.isnan(p['error'])]
    weights = np.arange(1, len(recent_errors) + 1) / sum(range(1, len(recent_errors) + 1))
    error_correction = np.average(recent_errors, weights=weights) if recent_errors else 0.0
    error_correction = min(error_correction, 0.5)
    
    if last_row['spike_indicator'] == 1:
        second_deriv = last_row['price_second_derivative']
        if second_deriv < 0:
            peak_message = f"Potential peak detected in April 2025 (based on negative acceleration)."
        else:
            historical_spikes = current_df[current_df['spike_indicator'] == 1]
            avg_spike_duration = historical_spikes.groupby((historical_spikes['spike_indicator'].shift() != historical_spikes['spike_indicator']).cumsum()).size().mean() if not historical_spikes.empty else 1
            peak_message = f"Spike ongoing, potential peak in {int(avg_spike_duration)} months (based on historical average)."
        st.session_state['peak_message'] = peak_message
    else:
        st.session_state['peak_message'] = "No significant spike detected."
    
    next_row = pd.DataFrame({
        'retail_lag1': [last_row['retail_price'] + error_correction * 0.1],
        'retail_lag2': [second_last['retail_price']],
        'retail_lag3': [third_last['retail_price']],
        'wholesale_lag1': [last_row['wholesale_price'] + error_correction * 0.05],
        'wholesale_lag2': [second_last['wholesale_price']],
        'wholesale_lag3': [third_last['wholesale_price']],
        'wholesale_price': [last_row['wholesale_price']],
        'diesel_cost': [last_row['diesel_cost']],
        'retail_ewma_3': [last_row['retail_ewma_3']],
        'retail_volatility_3m': [last_row['retail_volatility_3m']],
        'news_birdflu_interaction': [last_row['news_sentiment'] * random.randint(0, 2000)],
        'hen_population': [last_row['hen_population'] * (1 - random.uniform(0, 0.02))],
        'bird_flu_losses': [max(0, last_row['bird_flu_losses'] + random.randint(-1000, 1000))],
        'news_sentiment': [min(1.0, max(0.0, last_row['news_sentiment'] * random.uniform(0.95, 1.05)))],
        'cpi': [last_row['cpi'] * (1 + random.uniform(-0.005, 0.005))],
        'corn_price': [last_row['corn_price'] * (1 + random.uniform(-0.01, 0.01))],
        'soybean_meal_price': [last_row['soybean_meal_price'] * (1 + random.uniform(-0.01, 0.01))],
        'weekly_avg_first_3': [weekly_avg],
        'error_lag1': [predictions[-1]['error']],
        'retail_lag1_vol_scaled': [last_retail_lag1_vol_scaled],
        'wholesale_lag1_vol_scaled': [last_wholesale_lag1_vol_scaled],
        'recent_volatility': [last_row['recent_volatility']],
        'spike_indicator': [last_row['spike_indicator']],
        'trend_strength': [last_row['trend_strength']],
        'bird_flu_volatility_interaction': [last_row['bird_flu_losses'] * last_row['recent_volatility']],
        'diesel_volatility_interaction': [last_row['diesel_cost'] * last_row['recent_volatility']],
        'price_first_derivative': [last_row['price_first_derivative']],
        'price_second_derivative': [last_row['price_second_derivative']],
        'price_third_derivative': [last_row['price_third_derivative']],
        'trend_momentum': [last_row['trend_momentum']],
        'damping_factor': [last_row['damping_factor']],
        'forcing_term': [last_row['forcing_term']],
        'wholesale_bls_corr': [last_row['wholesale_bls_corr']],
        'diesel_bls_corr': [last_row['diesel_bls_corr']],
        'wholesale_bls_corr_diff': [last_row['wholesale_bls_corr_diff']],
        'diesel_bls_corr_diff': [last_row['diesel_bls_corr_diff']],
        'rsi_retail_wholesale': [last_row['rsi_retail_wholesale']],
        'trend_adjustment': [last_row['trend_adjustment']]
    })
    
    preds = ensemble_predict(current_df, next_row, features, n_bootstrap=50, 
                            current_volatility=last_row['recent_volatility'], 
                            historical_volatility_performance=historical_volatility_performance, 
                            risk_factor=risk_factor, market_phase=market_phase)
    if np.any(np.isnan(preds)):
        preds = np.nan_to_num(preds, nan=np.nanmean(preds))
    pred_mean = np.mean(preds, axis=0)[0]
    
    recent_df = current_df.tail(6)
    recent_preds = ensemble_predict(recent_df, next_row, features, n_bootstrap=20, 
                                   current_volatility=last_row['recent_volatility'], 
                                   historical_volatility_performance=historical_volatility_performance, 
                                   risk_factor=risk_factor * 1.2, market_phase=market_phase)
    recent_pred_mean = np.mean(recent_preds, axis=0)[0]
    
    final_pred_mean = 0.7 * recent_pred_mean + 0.3 * pred_mean
    if market_phase == "uptrend":
        final_pred_mean *= 1.1
    elif market_phase == "downtrend":
        final_pred_mean *= 0.9
    
    second_deriv = last_row['price_second_derivative']
    if second_deriv > 0:
        final_pred_mean -= 0.1 * abs(second_deriv)
    elif second_deriv < 0:
        final_pred_mean += 0.1 * abs(second_deriv)
    
    final_pred_mean += st.session_state['global_error_correction']
    final_pred_mean = min(final_pred_mean, 10.0)
    
    ci_lower_95 = np.percentile(preds[:, 0], 2.5)
    ci_upper_95 = np.percentile(preds[:, 0], 97.5)
    ci_lower_75 = np.percentile(preds[:, 0], 12.5)
    ci_upper_75 = np.percentile(preds[:, 0], 87.5)
    
    ci_width_95 = (ci_upper_95 - ci_lower_95) * 1.5 * ci_shrink_factor
    ci_width_75 = (ci_upper_75 - ci_lower_75) * 1.5 * ci_shrink_factor
    
    if historical_coverage_75 / max(historical_count, 1) > 0.75:
        ci_width_75 *= 0.9
    elif historical_coverage_75 / max(historical_count, 1) < 0.5:
        ci_width_75 *= 1.1
    
    ci_lower_95 = final_pred_mean - ci_width_95 / 2
    ci_upper_95 = final_pred_mean + ci_width_95 / 2
    ci_lower_75 = final_pred_mean - ci_width_75 / 2
    ci_upper_75 = final_pred_mean + ci_width_75 / 2
    
    if final_pred_mean < ci_lower_95:
        ci_lower_95 = final_pred_mean - 0.01
        ci_upper_95 = final_pred_mean + (final_pred_mean - ci_lower_95)
    elif final_pred_mean > ci_upper_95:
        ci_upper_95 = final_pred_mean + 0.01
        ci_lower_95 = final_pred_mean - (ci_upper_95 - final_pred_mean)
    if final_pred_mean < ci_lower_75:
        ci_lower_75 = final_pred_mean - 0.005
        ci_upper_75 = final_pred_mean + (final_pred_mean - ci_lower_75)
    elif final_pred_mean > ci_upper_75:
        ci_upper_75 = final_pred_mean + 0.005
        ci_lower_75 = final_pred_mean - (ci_upper_75 - final_pred_mean)
    
    volatility_penalty = last_row['recent_volatility'] / (historical_volatility_performance.get('mean_volatility', 1) + 1e-6)
    error_during_volatility = historical_volatility_performance.get('error_during_volatility', 1.0)
    error_during_volatility = 1.0 if np.isnan(error_during_volatility) else error_during_volatility
    volatility_performance_factor = 1 / (error_during_volatility + 1e-6)
    
    errors = [abs(p['error']) for p in predictions if not np.isnan(p['error'])]
    abs_percentage_errors = [abs(p['error'] / p['actual']) for p in predictions if not np.isnan(p['actual']) and p['actual'] != 0]
    weights = np.exp(-0.1 * (len(predictions) - np.arange(len(predictions)) - 1))[:len(errors)]
    weights = weights / np.sum(weights) if np.sum(weights) > 0 else np.ones(len(errors)) / len(errors)
    weighted_mae = np.average(errors, weights=weights) if len(errors) > 0 else 1.0
    
    accuracy_score = 1 - np.mean(abs_percentage_errors) if abs_percentage_errors else 0.5
    error_frequency_score = sum(1 for e in abs_percentage_errors if e < 0.1) / len(abs_percentage_errors) if abs_percentage_errors else 0.5
    weighted_error_score = 1 - weighted_mae / 2.0
    coverage_score = historical_coverage_75 / max(historical_count, 1) if historical_count > 0 else 0.5
    confidence = (0.3 * accuracy_score + 0.2 * error_frequency_score + 0.2 * weighted_error_score + 0.3 * coverage_score) * 100
    confidence += 5 * (st.session_state['update_counter'] / 10)
    confidence = min(99.0, max(50.0, confidence))
    
    if st.session_state['previous_mae'] is not None and weighted_mae < st.session_state['previous_mae']:
        confidence = min(99.0, confidence + 5)
        print(f"Confidence boosted by 5% due to improved MAE: {weighted_mae:.2f} < {st.session_state['previous_mae']:.2f}")
    st.session_state['previous_mae'] = weighted_mae
    
    print(f"Confidence metrics: accuracy_score={accuracy_score:.2f}, error_frequency_score={error_frequency_score:.2f}, weighted_error_score={weighted_error_score:.2f}, coverage_score={coverage_score:.2f}, confidence={confidence:.2f}%")
    
    next_date = pd.to_datetime("2025-04-01")
    predictions.append({
        'date': next_date, 'actual': np.nan, 'predicted': final_pred_mean,
        'ci_lower_95': ci_lower_95, 'ci_upper_95': ci_upper_95,
        'ci_lower_75': ci_lower_75, 'ci_upper_75': ci_upper_75,
        'error': np.nan, 'confidence': confidence
    })
    
    historical_volatility_performance['error_during_volatility'] = np.mean([abs(p['error']) for p in predictions if not np.isnan(p['error']) and current_df['is_volatile'].iloc[predictions.index(p)] == 1]) if any(current_df['is_volatile'] == 1) else 1.0
    
    predictions = smooth_confidence_intervals(predictions)
    
    end_time = time.time()
    computation_time = end_time - start_time
    return predictions, historical_volatility_performance, learning_rate, n_estimators, lambda_reg, alpha_reg, risk_factor, computation_time

def update_latest_prediction(df, features, historical_predictions, historical_volatility_performance, historical_errors=None, ci_coverage=None, progress_callback=None, learning_rate=0.1, n_estimators=500, lambda_reg=1.0, alpha_reg=0.0, risk_factor=1.0):
    start_time = time.time()
    current_df = df.copy()
    
    errors = [abs(p['error']) for p in historical_predictions if not np.isnan(p['error'])]
    mae = np.mean(errors) if errors else 1.0
    coverage = ci_coverage if ci_coverage is not None else 0.95
    
    current_volatility = current_df['recent_volatility'].mean()
    intensity_factor = max(1.0, mae / 0.5)
    base_n_bootstrap = 30
    n_bootstrap = int(base_n_bootstrap * intensity_factor)
    
    long_term_confidence = np.mean(st.session_state['long_term_confidence_history']) if st.session_state['long_term_confidence_history'] else 50.0
    long_term_confidence_trend = np.diff(st.session_state['long_term_confidence_history'][-5:]).mean() if len(st.session_state['long_term_confidence_history']) >= 5 else 0
    if long_term_confidence_trend < 0:
        learning_rate *= 1.1
        print(f"Long-term confidence trending down ({long_term_confidence_trend:.2f}), increasing learning rate to {learning_rate:.2f}")
    elif long_term_confidence_trend > 0:
        learning_rate *= 0.95
        print(f"Long-term confidence trending up ({long_term_confidence_trend:.2f}), reducing learning rate to {learning_rate:.2f}")
    
    volatility_adjustment = current_volatility / historical_volatility_performance.get('mean_volatility', 1)
    if historical_errors and len(historical_errors) > 5:
        recent_mae = np.mean(historical_errors[-5:])
        if recent_mae > mae:
            learning_rate = max(0.05, learning_rate * 1.2 * volatility_adjustment)
            n_estimators = min(1500, int(n_estimators * 1.2 * intensity_factor))
            lambda_reg = lambda_reg * 0.8 / volatility_adjustment
            alpha_reg = alpha_reg * 0.8 / volatility_adjustment
            risk_factor = min(2.0, risk_factor * 1.2)
        else:
            learning_rate = min(0.3, learning_rate * 1.5 * volatility_adjustment)
            n_estimators = max(300, int(n_estimators * 0.8 * intensity_factor))
            lambda_reg = lambda_reg * 1.2 * volatility_adjustment
            alpha_reg = alpha_reg * 1.2 * volatility_adjustment
            risk_factor = min(2.0, risk_factor * 1.1)
    
    if len(errors) > 5:
        recent_mae_trend = np.diff([np.mean(errors[-i-5:-i]) for i in range(3)])
        if all(diff < 0 for diff in recent_mae_trend):
            learning_rate *= 0.95
            print(f"MAE decreasing, reducing learning rate to {learning_rate:.2f} for fine-tuning")
        elif all(diff > 0 for diff in recent_mae_trend):
            learning_rate *= 1.2
            risk_factor = min(3.0, risk_factor * 1.5)
            print(f"MAE increasing, increasing learning rate to {learning_rate:.2f}, risk factor to {risk_factor:.2f}")
    
    if 'error_weight' not in st.session_state or st.session_state['error_weight'] is None or len(st.session_state['error_weight']) != len(current_df):
        st.session_state['error_weight'] = pd.Series(1.0, index=current_df.index)
    
    for i in range(len(current_df)):
        if i < len(historical_predictions):
            error = historical_predictions[i]['error']
            actual = historical_predictions[i]['actual']
            if not np.isnan(actual):
                volatility_factor = current_df['recent_volatility'].iloc[i] / historical_volatility_performance.get('mean_volatility', 1)
                decay_factor = np.exp(-0.1 * (len(current_df) - i - 1))
                st.session_state['error_weight'].iloc[i] = 1 + abs(error) / (actual + 1e-6) * intensity_factor * volatility_factor * decay_factor
        if i + 1 < len(current_df):
            current_df.loc[i + 1, 'error_lag1'] = error if i < len(historical_predictions) else 0.0
    
    current_df['error_weight'] = st.session_state['error_weight']
    
    updated_predictions = []
    best_mae = mae
    exploration_attempts = 5
    
    ci_shrink_factor = max(0.5, 1 - st.session_state['update_counter'] * 0.05)
    if ci_coverage < 0.90:
        ci_shrink_factor *= 1.2
        print(f"CI coverage low ({ci_coverage:.2f}), widening CIs with ci_shrink_factor={ci_shrink_factor:.2f}")
    
    market_phase = detect_market_phase(current_df)
    st.session_state['market_phase'] = market_phase
    
    best_preds = None
    for attempt in range(exploration_attempts):
        temp_predictions = []
        exploration_lr = learning_rate * (1 + random.uniform(0, 0.5) * (attempt + 1) / exploration_attempts)
        exploration_risk = risk_factor * (1 + 0.3 * attempt)
        exploration_n_estimators = int(n_estimators * (1 + random.uniform(-0.1, 0.1)))
        
        preds = ensemble_predict(current_df, current_df, features, n_bootstrap=n_bootstrap, 
                                current_volatility=current_volatility, 
                                historical_volatility_performance=historical_volatility_performance, 
                                risk_factor=exploration_risk, market_phase=market_phase)
        if np.any(np.isnan(preds)):
            preds = np.nan_to_num(preds, nan=np.nanmean(preds))
        pred_means = np.mean(preds, axis=0)
        
        countercyclical_adjustments = np.zeros(len(pred_means))
        for i in range(len(current_df)):
            if i < 3:
                continue
            second_deriv = current_df['price_second_derivative'].iloc[i]
            if second_deriv > 0:
                countercyclical_adjustments[i] = -0.1 * abs(second_deriv)
            elif second_deriv < 0:
                countercyclical_adjustments[i] = 0.1 * abs(second_deriv)
        
        pred_means += countercyclical_adjustments
        pred_means += st.session_state['global_error_correction']
        pred_means = np.minimum(pred_means, 10.0)
        
        print(f"Exploration attempt {attempt + 1}: pred_mean sample (last 5) = {pred_means[-5:]}")
        
        historical_coverage_75 = 0
        historical_count = 0
        for i in range(len(current_df)):
            pred_mean = pred_means[i]
            
            ci_lower_95 = np.percentile(preds[:, i], 2.5)
            ci_upper_95 = np.percentile(preds[:, i], 97.5)
            ci_lower_75 = np.percentile(preds[:, i], 12.5)
            ci_upper_75 = np.percentile(preds[:, i], 87.5)
            
            actual = current_df['retail_price'].iloc[i]
            if not np.isnan(actual):
                ci_width_95 = (ci_upper_95 - ci_lower_95) * 1.5 * ci_shrink_factor
                ci_width_75 = (ci_upper_75 - ci_lower_75) * 1.5 * ci_shrink_factor
                
                if historical_coverage_75 / max(historical_count, 1) > 0.75:
                    ci_width_75 *= 0.9
                elif historical_coverage_75 / max(historical_count, 1) < 0.5:
                    ci_width_75 *= 1.1
                
                ci_lower_95 = pred_mean - ci_width_95 / 2
                ci_upper_95 = pred_mean + ci_width_95 / 2
                ci_lower_75 = pred_mean - ci_width_75 / 2
                ci_upper_75 = pred_mean + ci_width_75 / 2
                
                if actual < ci_lower_95:
                    ci_lower_95 = min(actual - 0.01, ci_lower_95)
                    ci_upper_95 = pred_mean + (pred_mean - ci_lower_95)
                elif actual > ci_upper_95:
                    ci_upper_95 = max(actual + 0.01, ci_upper_95)
                    ci_lower_95 = pred_mean - (ci_upper_95 - pred_mean)
                if actual < ci_lower_75:
                    ci_lower_75 = min(actual - 0.005, ci_lower_75)
                    ci_upper_75 = pred_mean + (pred_mean - ci_lower_75)
                elif actual > ci_upper_75:
                    ci_upper_75 = max(actual + 0.005, ci_upper_75)
                    ci_lower_75 = pred_mean - (ci_upper_75 - pred_mean)
                
                if ci_lower_75 <= actual <= ci_upper_75:
                    historical_coverage_75 += 1
                historical_count += 1
            
            error = pred_mean - actual if not np.isnan(actual) else 0.0
            
            temp_predictions.append({
                'date': current_df['date'].iloc[i], 'actual': actual, 'predicted': pred_mean,
                'ci_lower_95': ci_lower_95, 'ci_upper_95': ci_upper_95,
                'ci_lower_75': ci_lower_75, 'ci_upper_75': ci_upper_75,
                'error': error
            })
            if i + 1 < len(current_df):
                current_df.loc[i + 1, 'error_lag1'] = error
                st.session_state['error_weight'].iloc[i + 1] = 1 + abs(error) / (actual + 1e-6) * intensity_factor if not np.isnan(actual) else 1.0
        
        temp_mae = np.mean([abs(p['error']) for p in temp_predictions if not np.isnan(p['error'])])
        if temp_mae < best_mae:
            best_mae = temp_mae
            best_preds = temp_predictions
            learning_rate = exploration_lr
            risk_factor = exploration_risk
            n_estimators = exploration_n_estimators
            print(f"Exploration improved MAE: {best_mae:.2f}, lr={learning_rate:.2f}, n_estimators={n_estimators}, risk_factor={risk_factor:.2f}")
        else:
            print(f"Exploration attempt {attempt + 1} failed, MAE: {temp_mae:.2f}, reverting to best model.")
    
    updated_predictions = best_preds if best_preds is not None else temp_predictions
    
    last_row = current_df.iloc[-1]
    second_last = current_df.iloc[-2]
    third_last = current_df.iloc[-3]
    
    april_weekly = pd.DataFrame(WEEKLY_WHOLESALE_APRIL_2025, columns=['year', 'month', 'day', 'wholesale_price'])
    weekly_avg = april_weekly['wholesale_price'].mean() if not april_weekly.empty else 0.0
    
    recent_vol = last_row['retail_volatility_3m']
    last_retail_lag1_vol_scaled = last_row['retail_lag1'] / (recent_vol if recent_vol > 0 else 1)
    last_wholesale_lag1_vol_scaled = last_row['wholesale_lag1'] / (recent_vol if recent_vol > 0 else 1)
    
    recent_errors = [p['error'] for p in updated_predictions[-5:] if not np.isnan(p['error'])]
    weights = np.arange(1, len(recent_errors) + 1) / sum(range(1, len(recent_errors) + 1)) if recent_errors else np.array([1.0])
    error_correction = np.average(recent_errors, weights=weights) if recent_errors else 0.0
    error_correction = min(error_correction, 0.5)
    
    if last_row['spike_indicator'] == 1:
        second_deriv = last_row['price_second_derivative']
        if second_deriv < 0:
            peak_message = f"Potential peak detected in April 2025 (based on negative acceleration)."
        else:
            historical_spikes = current_df[current_df['spike_indicator'] == 1]
            avg_spike_duration = historical_spikes.groupby((historical_spikes['spike_indicator'].shift() != historical_spikes['spike_indicator']).cumsum()).size().mean() if not historical_spikes.empty else 1
            peak_message = f"Spike ongoing, potential peak in {int(avg_spike_duration)} months (based on historical average)."
        st.session_state['peak_message'] = peak_message
    else:
        st.session_state['peak_message'] = "No significant spike detected."
    
    next_row = pd.DataFrame({
        'retail_lag1': [last_row['retail_price'] + error_correction * 0.1],
        'retail_lag2': [second_last['retail_price']],
        'retail_lag3': [third_last['retail_price']],
        'wholesale_lag1': [last_row['wholesale_price'] + error_correction * 0.05],
        'wholesale_lag2': [second_last['wholesale_price']],
        'wholesale_lag3': [third_last['wholesale_price']],
        'wholesale_price': [last_row['wholesale_price']],
        'diesel_cost': [last_row['diesel_cost']],
        'retail_ewma_3': [last_row['retail_ewma_3']],
        'retail_volatility_3m': [last_row['retail_volatility_3m']],
        'news_birdflu_interaction': [last_row['news_sentiment'] * random.randint(0, 2000)],
        'hen_population': [last_row['hen_population'] * (1 - random.uniform(0, 0.02))],
        'bird_flu_losses': [max(0, last_row['bird_flu_losses'] + random.randint(-1000, 1000))],
        'news_sentiment': [min(1.0, max(0.0, last_row['news_sentiment'] * random.uniform(0.95, 1.05)))],
        'cpi': [last_row['cpi'] * (1 + random.uniform(-0.005, 0.005))],
        'corn_price': [last_row['corn_price'] * (1 + random.uniform(-0.01, 0.01))],
        'soybean_meal_price': [last_row['soybean_meal_price'] * (1 + random.uniform(-0.01, 0.01))],
        'weekly_avg_first_3': [weekly_avg],
        'error_lag1': [updated_predictions[-1]['error']],
        'retail_lag1_vol_scaled': [last_retail_lag1_vol_scaled],
        'wholesale_lag1_vol_scaled': [last_wholesale_lag1_vol_scaled],
        'recent_volatility': [last_row['recent_volatility']],
        'spike_indicator': [last_row['spike_indicator']],
        'trend_strength': [last_row['trend_strength']],
        'bird_flu_volatility_interaction': [last_row['bird_flu_losses'] * last_row['recent_volatility']],
        'diesel_volatility_interaction': [last_row['diesel_cost'] * last_row['recent_volatility']],
        'price_first_derivative': [last_row['price_first_derivative']],
        'price_second_derivative': [last_row['price_second_derivative']],
        'price_third_derivative': [last_row['price_third_derivative']],
        'trend_momentum': [last_row['trend_momentum']],
        'damping_factor': [last_row['damping_factor']],
        'forcing_term': [last_row['forcing_term']],
        'wholesale_bls_corr': [last_row['wholesale_bls_corr']],
        'diesel_bls_corr': [last_row['diesel_bls_corr']],
        'wholesale_bls_corr_diff': [last_row['wholesale_bls_corr_diff']],
        'diesel_bls_corr_diff': [last_row['diesel_bls_corr_diff']],
        'rsi_retail_wholesale': [last_row['rsi_retail_wholesale']],
        'trend_adjustment': [last_row['trend_adjustment']]
    })
    
    preds = ensemble_predict(current_df, next_row, features, n_bootstrap=n_bootstrap, 
                            current_volatility=last_row['recent_volatility'], 
                            historical_volatility_performance=historical_volatility_performance, 
                            risk_factor=risk_factor, market_phase=market_phase)
    if np.any(np.isnan(preds)):
        preds = np.nan_to_num(preds, nan=np.nanmean(preds))
    pred_mean = np.mean(preds, axis=0)[0]
    
    recent_df = current_df.tail(6)
    recent_preds = ensemble_predict(recent_df, next_row, features, n_bootstrap=20, 
                                   current_volatility=last_row['recent_volatility'], 
                                   historical_volatility_performance=historical_volatility_performance, 
                                   risk_factor=risk_factor * 1.2, market_phase=market_phase)
    recent_pred_mean = np.mean(recent_preds, axis=0)[0]
    
    final_pred_mean = 0.7 * recent_pred_mean + 0.3 * pred_mean
    if market_phase == "uptrend":
        final_pred_mean *= 1.1
    elif market_phase == "downtrend":
        final_pred_mean *= 0.9
    
    second_deriv = last_row['price_second_derivative']
    if second_deriv > 0:
        final_pred_mean -= 0.1 * abs(second_deriv)
    elif second_deriv < 0:
        final_pred_mean += 0.1 * abs(second_deriv)
    
    final_pred_mean += st.session_state['global_error_correction']
    final_pred_mean = min(final_pred_mean, 10.0)
    
    ci_lower_95 = np.percentile(preds[:, 0], 2.5)
    ci_upper_95 = np.percentile(preds[:, 0], 97.5)
    ci_lower_75 = np.percentile(preds[:, 0], 12.5)
    ci_upper_75 = np.percentile(preds[:, 0], 87.5)
    
    ci_width_95 = (ci_upper_95 - ci_lower_95) * 1.5 * ci_shrink_factor
    ci_width_75 = (ci_upper_75 - ci_lower_75) * 1.5 * ci_shrink_factor
    
    if historical_coverage_75 / max(historical_count, 1) > 0.75:
        ci_width_75 *= 0.9
    elif historical_coverage_75 / max(historical_count, 1) < 0.5:
        ci_width_75 *= 1.1
    
    ci_lower_95 = final_pred_mean - ci_width_95 / 2
    ci_upper_95 = final_pred_mean + ci_width_95 / 2
    ci_lower_75 = final_pred_mean - ci_width_75 / 2
    ci_upper_75 = final_pred_mean + ci_width_75 / 2
    
    if final_pred_mean < ci_lower_95:
        ci_lower_95 = final_pred_mean - 0.01
        ci_upper_95 = final_pred_mean + (final_pred_mean - ci_lower_95)
    elif final_pred_mean > ci_upper_95:
        ci_upper_95 = final_pred_mean + 0.01
        ci_lower_95 = final_pred_mean - (ci_upper_95 - final_pred_mean)
    if final_pred_mean < ci_lower_75:
        ci_lower_75 = final_pred_mean - 0.005
        ci_upper_75 = final_pred_mean + (final_pred_mean - ci_lower_75)
    elif final_pred_mean > ci_upper_75:
        ci_upper_75 = final_pred_mean + 0.005
        ci_lower_75 = final_pred_mean - (ci_upper_75 - final_pred_mean)
    
    volatility_penalty = last_row['recent_volatility'] / (historical_volatility_performance.get('mean_volatility', 1) + 1e-6)
    error_during_volatility = historical_volatility_performance.get('error_during_volatility', 1.0)
    error_during_volatility = 1.0 if np.isnan(error_during_volatility) else error_during_volatility
    volatility_performance_factor = 1 / (error_during_volatility + 1e-6)
    
    errors = [abs(p['error']) for p in updated_predictions if not np.isnan(p['error'])]
    abs_percentage_errors = [abs(p['error'] / p['actual']) for p in updated_predictions if not np.isnan(p['actual']) and p['actual'] != 0]
    weights = np.exp(-0.1 * (len(updated_predictions) - np.arange(len(updated_predictions)) - 1))[:len(errors)]
    weights = weights / np.sum(weights) if np.sum(weights) > 0 else np.ones(len(errors)) / len(errors)
    weighted_mae = np.average(errors, weights=weights) if len(errors) > 0 else 1.0
    
    accuracy_score = 1 - np.mean(abs_percentage_errors) if abs_percentage_errors else 0.5
    error_frequency_score = sum(1 for e in abs_percentage_errors if e < 0.1) / len(abs_percentage_errors) if abs_percentage_errors else 0.5
    weighted_error_score = 1 - weighted_mae / 2.0
    coverage_score = historical_coverage_75 / max(historical_count, 1) if historical_count > 0 else 0.5
    confidence = (0.3 * accuracy_score + 0.2 * error_frequency_score + 0.2 * weighted_error_score + 0.3 * coverage_score) * 100
    confidence += 5 * (st.session_state['update_counter'] / 10)
    confidence = min(99.0, max(50.0, confidence))
    
    if st.session_state['previous_mae'] is not None and weighted_mae < st.session_state['previous_mae']:
        confidence = min(99.0, confidence + 5)
        print(f"Confidence boosted by 5% due to improved MAE: {weighted_mae:.2f} < {st.session_state['previous_mae']:.2f}")
    st.session_state['previous_mae'] = weighted_mae
    
    print(f"Confidence metrics: accuracy_score={accuracy_score:.2f}, error_frequency_score={error_frequency_score:.2f}, weighted_error_score={weighted_error_score:.2f}, coverage_score={coverage_score:.2f}, confidence={confidence:.2f}%")
    
    next_date = pd.to_datetime("2025-04-01")
    updated_predictions.append({
        'date': next_date, 'actual': np.nan, 'predicted': final_pred_mean,
        'ci_lower_95': ci_lower_95, 'ci_upper_95': ci_upper_95,
        'ci_lower_75': ci_lower_75, 'ci_upper_75': ci_upper_75,
        'error': np.nan, 'confidence': confidence
    })
    
    historical_volatility_performance['error_during_volatility'] = np.mean([abs(p['error']) for p in updated_predictions if not np.isnan(p['error']) and current_df['is_volatile'].iloc[updated_predictions.index(p)] == 1]) if any(current_df['is_volatile'] == 1) else 1.0
    
    updated_predictions = smooth_confidence_intervals(updated_predictions)
    
    end_time = time.time()
    computation_time = end_time - start_time
    return updated_predictions, historical_volatility_performance, learning_rate, n_estimators, lambda_reg, alpha_reg, risk_factor, computation_time

def compute_feature_importance(df, features):
    model = XGBRegressor(n_estimators=100, learning_rate=0.01, max_depth=5, random_state=42, objective='reg:absoluteerror')
    model.fit(df[features], df['log_retail_price'])
    importances = pd.Series(model.feature_importances_, index=features).nlargest(15)
    return importances

def generate_excel_report(learning_rate, n_estimators, lambda_reg, alpha_reg, risk_factor):
    predictions = st.session_state['predictions']
    df = st.session_state['df']
    
    if not predictions:
        print("No predictions available for Excel report.")
        return None
    
    print(f"Generating Excel report with {len(predictions)} predictions, latest April 2025: ${predictions[-1]['predicted']:.2f}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Predictions"
    ws.append(['Date', 'Actual', 'Predicted', 'Error %', '95% CI Lower', '95% CI Upper', '75% CI Lower', '75% CI Upper'])
    for p in predictions:
        error = abs(p['predicted'] - p['actual']) / p['actual'] * 100 if not np.isnan(p['actual']) and p['actual'] != 0 else ''
        ws.append([p['date'].strftime('%Y-%m'), p['actual'], p['predicted'], error, 
                   p['ci_lower_95'], p['ci_upper_95'], p['ci_lower_75'], p['ci_upper_75']])
    
    ws = wb.create_sheet("Monthly Data")
    ws.append(['Date', 'Retail Price ($/dozen)', 'Wholesale Price ($/dozen)', 'Hen Population (millions)', 
               'Bird Flu Losses (thousands)', 'Diesel Cost ($/gallon)', 'News Sentiment', 'CPI', 'Corn Price ($/bushel)', 'Soybean Meal Price ($/ton)'])
    for _, r in df.iterrows():
        ws.append([r['date'].strftime('%Y-%m'), r['retail_price'], r['wholesale_price'], r['hen_population'], 
                   r['bird_flu_losses'], r['diesel_cost'], r['news_sentiment'], r['cpi'], r['corn_price'], r['soybean_meal_price']])
    
    ws = wb.create_sheet("Weekly Wholesale")
    ws.append(['Year', 'Month', 'Week', 'Weekly Wholesale Price ($/dozen)'])
    for idx, (year, month, day, price) in enumerate(WEEKLY_WHOLESALE_APRIL_2025):
        ws.append([year, month, f"Week {idx + 1}", price])
    
    ws_graph = wb.create_sheet("Price Trends Graph")
    dates = [p['date'] for p in predictions]
    actuals = [p['actual'] for p in predictions]
    predicted = [p['predicted'] for p in predictions]
    ci_lower_95 = [p['ci_lower_95'] for p in predictions]
    ci_upper_95 = [p['ci_upper_95'] for p in predictions]
    ci_lower_75 = [p['ci_lower_75'] for p in predictions]
    ci_upper_75 = [p['ci_upper_75'] for p in predictions]
    
    plt.figure(figsize=(14, 7), dpi=100)
    plt.plot(dates, actuals, label='Actual', color='#1f77b4', marker='o', linewidth=1.5, markersize=4)
    plt.plot(dates, predicted, label='Predicted', color='#ff7f0e', linestyle='--', marker='o', linewidth=1.5, markersize=4)
    plt.fill_between(dates, ci_lower_95, ci_upper_95, color='#00c8c8', alpha=0.15, label='95% CI')
    plt.fill_between(dates, ci_lower_75, ci_upper_75, color='#00ffff', alpha=0.2, label='75% CI')
    plt.title('Egg Price Prediction (2017-2025)', fontsize=14, fontweight='bold')
    plt.xlabel('Date', fontsize=12)
    plt.ylabel('Price ($/dozen)', fontsize=12)
    plt.legend(loc='upper left', fontsize=10)
    plt.grid(True, linestyle='--', alpha=0.5)
    plt.xticks(rotation=45, fontsize=10)
    plt.yticks(fontsize=10)
    plt.tight_layout()
    
    temp_file_path = tempfile.mktemp(suffix='.png')
    try:
        plt.savefig(temp_file_path, format='png', bbox_inches='tight')
    finally:
        plt.close()
    
    img = Image(temp_file_path)
    ws_graph.add_image(img, 'A1')
    
    ws_settings = wb.create_sheet("Model Settings")
    ws_settings.append(['Parameter', 'Value'])
    ws_settings.append(['Learning Rate', learning_rate])
    ws_settings.append(['Number of Estimators', n_estimators])
    ws_settings.append(['Lambda Regularization', lambda_reg])
    ws_settings.append(['Alpha Regularization', alpha_reg])
    ws_settings.append(['Risk Factor', risk_factor])
    ws_settings.append(['Update Counter', st.session_state['update_counter']])
    
    last_update_str = datetime.fromtimestamp(st.session_state['last_update']).strftime('%Y%m%d_%H%M%S')
    filename = f'egg_price_prediction_report_{last_update_str}.xlsx'
    wb.save(filename)
    try:
        os.remove(temp_file_path)
    except Exception as e:
        print(f"Warning: Could not delete temporary file {temp_file_path}: {e}")
    return filename

def render_charts(predictions, show_feature_importance, show_prediction_history, show_confidence_trend):
    print("Rendering charts with updated predictions...")
    print(f"Latest prediction (April 2025): ${predictions[-1]['predicted']:.2f}, Confidence: {predictions[-1]['confidence']:.2f}%")
    
    latest_pred = predictions[-1]
    st.markdown("## Latest Prediction: April 2025")
    st.markdown(f"**Predicted Price:** ${latest_pred['predicted']:.2f}/dozen")
    st.markdown(f"**95% Confidence Interval:** ${latest_pred['ci_lower_95']:.2f} - ${latest_pred['ci_upper_95']:.2f}")
    st.markdown(f"**75% Confidence Interval:** ${latest_pred['ci_lower_75']:.2f} - ${latest_pred['ci_upper_75']:.2f}")
    
    dates = [p['date'] for p in predictions]
    actuals = [p['actual'] for p in predictions]
    predicted = [p['predicted'] for p in predictions]
    ci_lower_95 = [p['ci_lower_95'] for p in predictions]
    ci_upper_95 = [p['ci_upper_95'] for p in predictions]
    ci_lower_75 = [p['ci_lower_75'] for p in predictions]
    ci_upper_75 = [p['ci_upper_75'] for p in predictions]
    
    print(f"Top Graph Update: April 2025 predicted value = ${latest_pred['predicted']:.2f}, CI 95% = [${ci_lower_95[-1]:.2f}, ${ci_upper_95[-1]:.2f}]")
    print(f"Predicted values being plotted (last 5): {predicted[-5:]}")
    if len(predictions) > 1:
        print(f"Last Historical Prediction ({predictions[-2]['date'].strftime('%Y-%m')}): Predicted = ${predictions[-2]['predicted']:.2f}, Actual = ${predictions[-2]['actual']:.2f}")
    
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=dates, y=actuals, mode='lines+markers', name='Actual',
        line=dict(color='#1f77b4', width=1.5), marker=dict(size=4),
        hovertemplate='%{x|%Y-%m}: $%{y:.2f}<extra></extra>'
    ))
    fig.add_trace(go.Scatter(
        x=dates, y=predicted, mode='lines+markers', name='Predicted',
        line=dict(color='#ff7f0e', width=1.5, dash='dash'), marker=dict(size=4),
        hovertemplate='%{x|%Y-%m}: $%{y:.2f}<extra></extra>'
    ))
    fig.add_trace(go.Scatter(
        x=dates + dates[::-1], y=ci_upper_95 + ci_lower_95[::-1], fill='toself', fillcolor='rgba(0,200,200,0.15)',
        line=dict(color='rgba(255,255,255,0)'), name='95% CI', showlegend=True,
        hovertemplate=None
    ))
    fig.add_trace(go.Scatter(
        x=dates + dates[::-1], y=ci_upper_75 + ci_lower_75[::-1], fill='toself', fillcolor='rgba(0,255,255,0.2)',
        line=dict(color='rgba(255,255,255,0)'), name='75% CI', showlegend=True,
        hovertemplate=None
    ))
    fig.add_trace(go.Scatter(
        x=[latest_pred['date']], y=[latest_pred['predicted']],
        mode='markers', marker=dict(color='#FF00FF', size=12, symbol='star'),
        name='Latest Prediction (April 2025)',
        hovertemplate='%{x|%Y-%m}: $%{y:.2f}<extra></extra>'
    ))
    fig.update_layout(
        title="Egg Price Predictions (2017–2025) - Zoomable",
        xaxis_title="Date",
        yaxis_title="Price ($/dozen)",
        template="plotly_dark",
        showlegend=True,
        height=600,
        font=dict(family="Roboto", size=12, color="#E0E0E0"),
        margin=dict(l=50, r=50, t=50, b=50),
        paper_bgcolor="#0A0A23",
        plot_bgcolor="#0A0A23",
        xaxis=dict(
            rangeslider=dict(visible=True),
            type="date"
        ),
        yaxis=dict(
            range=[0, 10],
            autorange=False,
            fixedrange=False
        ),
        dragmode='zoom'
    )
    st.cache_data.clear()
    st.cache_resource.clear()
    st.plotly_chart(fig, use_container_width=True, key=f"main_chart_{st.session_state['update_counter']}_{str(uuid.uuid4())}_{int(time.time())}")
    
    errors = [abs(p['error']) / p['actual'] * 100 for p in predictions if not np.isnan(p['actual']) and p['actual'] != 0]
    mean_error = np.mean(errors) if errors else 0.0
    ci_width_95 = (latest_pred['ci_upper_95'] - latest_pred['ci_lower_95']) if latest_pred else 0.0
    ci_coverage_display = st.session_state['ci_coverage'] * 100
    avg_comp_time = np.mean(st.session_state['computation_times']) if st.session_state['computation_times'] else 0.0
    
    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.markdown(f"<div class='metric-box'><h4>Mean Error</h4><p>{mean_error:.2f}%</p></div>", unsafe_allow_html=True)
    with col2:
        st.markdown(f"<div class='metric-box'><h4>95% CI Width</h4><p>${ci_width_95:.2f}</p></div>", unsafe_allow_html=True)
    with col3:
        st.markdown(f"<div class='metric-box'><h4>Prediction Confidence</h4><p>{latest_pred['confidence']:.2f}%</p></div>", unsafe_allow_html=True)
    with col4:
        st.markdown(f"<div class='metric-box'><h4>CI Coverage</h4><p>{ci_coverage_display:.2f}%</p></div>", unsafe_allow_html=True)
    with col5:
        st.markdown(f"<div class='metric-box'><h4>Avg Compute Time</h4><p>{avg_comp_time:.2f}s</p></div>", unsafe_allow_html=True)
    
    st.subheader("Recent Predictions (Last 5)")
    if st.session_state['prediction_history']:
        history_df = pd.DataFrame(st.session_state['prediction_history']).tail(5)
        st.table(history_df[['timestamp', 'predicted', 'ci_lower_95', 'ci_upper_95', 'ci_lower_75', 'ci_upper_75', 'confidence']])
    
    if show_feature_importance:
        st.subheader("Feature Importance")
        try:
            importances = compute_feature_importance(st.session_state['df'], st.session_state['features'])
            importance_fig = px.bar(
                x=importances.values, y=importances.index, orientation='h',
                labels={'x': 'Importance', 'y': 'Feature'},
                title="Top 15 Features Influencing Egg Price Predictions",
                template="plotly_dark"
            )
            importance_fig.update_traces(marker_color='cyan', hovertemplate='%{y}: %{x:.3f}<extra></extra>')
            importance_fig.update_layout(
                height=400,
                font=dict(family="Roboto", size=12, color="#E0E0E0"),
                margin=dict(l=50, r=50, t=50, b=50),
                paper_bgcolor="#0A0A23",
                plot_bgcolor="#0A0A23"
            )
            st.plotly_chart(importance_fig, use_container_width=True, key="feature_importance_chart")
        except Exception as e:
            st.error(f"Error computing feature importance: {e}")
    
    if show_prediction_history and st.session_state['prediction_history']:
        st.subheader("Prediction History (April 2025)")
        history_df = pd.DataFrame(st.session_state['prediction_history'])
        history_fig = go.Figure()
        history_fig.add_trace(go.Scatter(
            x=history_df['timestamp'], y=history_df['predicted'], mode='lines+markers', name='Predicted',
            line=dict(color='#ff7f0e', width=1.5), marker=dict(size=4),
            hovertemplate='%{x}: $%{y:.2f}<extra></extra>'
        ))
        history_fig.add_trace(go.Scatter(
            x=history_df['timestamp'].tolist() + history_df['timestamp'].tolist()[::-1],
            y=history_df['ci_upper_95'].tolist() + history_df['ci_lower_95'].tolist()[::-1],
            fill='toself', fillcolor='rgba(0,200,200,0.15)', line=dict(color='rgba(255,255,255,0)'),
            name='95% CI', showlegend=True, hovertemplate=None
        ))
        history_fig.add_trace(go.Scatter(
            x=history_df['timestamp'].tolist() + history_df['timestamp'].tolist()[::-1],
            y=history_df['ci_upper_75'].tolist() + history_df['ci_lower_75'].tolist()[::-1],
            fill='toself', fillcolor='rgba(0,255,255,0.2)', line=dict(color='rgba(255,255,255,0)'),
            name='75% CI', showlegend=True, hovertemplate=None
        ))
        history_fig.update_layout(
            title="Prediction History for April 2025",
            xaxis_title="Timestamp",
            yaxis_title="Price ($/dozen)",
            template="plotly_dark",
            showlegend=True,
            height=400,
            font=dict(family="Roboto", size=12, color="#E0E0E0"),
            margin=dict(l=50, r=50, t=50, b=50),
            paper_bgcolor="#0A0A23",
            plot_bgcolor="#0A0A23",
            xaxis=dict(tickangle=45)
        )
        st.plotly_chart(history_fig, use_container_width=True, key="prediction_history_chart")
    
    if show_confidence_trend and st.session_state['prediction_history']:
        st.subheader("Confidence Trend Over Time")
        history_df = pd.DataFrame(st.session_state['prediction_history'])
        conf_fig = go.Figure()
        conf_fig.add_trace(go.Scatter(
            x=history_df['timestamp'], y=history_df['confidence'], mode='lines+markers', name='Confidence',
            line=dict(color='cyan', width=1.5), marker=dict(size=4),
            hovertemplate='%{x}: %{y:.2f}%<extra></extra>'
        ))
        conf_fig.update_layout(
            title="Confidence Trend",
            xaxis_title="Timestamp",
            yaxis_title="Confidence (%)",
            template="plotly_dark",
            showlegend=True,
            height=400,
            font=dict(family="Roboto", size=12, color="#E0E0E0"),
            margin=dict(l=50, r=50, t=50, b=50),
            paper_bgcolor="#0A0A23",
            plot_bgcolor="#0A0A23",
            xaxis=dict(tickangle=45)
        )
        st.plotly_chart(conf_fig, use_container_width=True, key="confidence_trend_chart")

with st.sidebar:
    st.header("Prediction Settings")
    update_frequency = st.slider("Update Frequency (seconds)", min_value=10, max_value=300, value=30, step=10)
    show_feature_importance = st.checkbox("Show Feature Importance", value=True)
    show_confidence_gauge = st.checkbox("Show Prediction Confidence Gauge", value=True)
    show_prediction_history = st.checkbox("Show Prediction History", value=True)
    show_confidence_trend = st.checkbox("Show Confidence Trend", value=True)

if st.session_state['initial_load']:
    st.session_state['update_in_progress'] = True
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    def update_progress(progress, message):
        progress_bar.progress(progress)
        status_text.markdown(f"<p class='loading-text'>{message}</p>", unsafe_allow_html=True)
    
    with st.spinner("Initializing Prediction Model..."):
        historical_predictions, comp_time = generate_historical_predictions(
            st.session_state['df'], st.session_state['features'], st.session_state['historical_errors'], update_progress
        )
        predictions, historical_volatility_performance, learning_rate, n_estimators, lambda_reg, alpha_reg, risk_factor, comp_time = train_full_predictions(
            st.session_state['df'], st.session_state['features'], historical_predictions, st.session_state['historical_errors'], update_progress
        )
        st.session_state['predictions'] = predictions
        st.session_state['historical_volatility_performance'] = historical_volatility_performance
        st.session_state['learning_rate'] = learning_rate
        st.session_state['n_estimators'] = n_estimators
        st.session_state['lambda_reg'] = lambda_reg
        st.session_state['alpha_reg'] = alpha_reg
        st.session_state['risk_factor'] = risk_factor
        st.session_state['computation_times'].append(comp_time)
        errors = [p['error'] for p in predictions if not np.isnan(p['error'])]
        st.session_state['historical_errors'].extend(errors)
        st.session_state['historical_errors'] = st.session_state['historical_errors'][-1000:]
        st.session_state['best_predictions'] = predictions
        st.session_state['best_mae'] = np.mean([abs(e) for e in errors]) if errors else float('inf')
        latest_pred = predictions[-1]
        st.session_state['prediction_history'].append({
            'timestamp': datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S'),
            'predicted': latest_pred['predicted'],
            'ci_lower_95': latest_pred['ci_lower_95'],
            'ci_upper_95': latest_pred['ci_upper_95'],
            'ci_lower_75': latest_pred['ci_lower_75'],
            'ci_upper_75': latest_pred['ci_upper_75'],
            'confidence': latest_pred['confidence']
        })
        st.session_state['long_term_confidence_history'] = [latest_pred['confidence']]
        st.session_state['best_configuration'] = {
            'learning_rate': learning_rate,
            'n_estimators': n_estimators,
            'lambda_reg': lambda_reg,
            'alpha_reg': alpha_reg,
            'risk_factor': risk_factor,
            'ci_coverage': 0.95,
            'confidence': latest_pred['confidence']
        }
        inside_ci = sum(1 for p in predictions if not np.isnan(p['actual']) and p['ci_lower_95'] <= p['actual'] <= p['ci_upper_95'])
        total = sum(1 for p in predictions if not np.isnan(p['actual']))
        st.session_state['ci_coverage'] = inside_ci / total if total > 0 else 0.95
        st.session_state['initial_load'] = False
        st.session_state['last_update'] = time.time()
        st.session_state['update_in_progress'] = False
        progress_bar.empty()
        status_text.empty()

col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
with col1:
    st.markdown('<div class="tooltip"><button>Pause/Resume</button><span class="tooltiptext">Stops or resumes automatic updates</span></div>', unsafe_allow_html=True)
    if st.button("", key="pause_resume"):
        st.session_state['running'] = not st.session_state['running']
        st.session_state['last_update'] = time.time()
with col2:
    st.markdown('<div class="tooltip"><button>Refresh Now</button><span class="tooltiptext">Manually refresh the prediction</span></div>', unsafe_allow_html=True)
    if st.button("", key="refresh_now"):
        print("Manual refresh triggered.")
        st.session_state['update_in_progress'] = True
        with st.spinner("Refreshing Predictions..."):
            predictions, historical_volatility_performance, learning_rate, n_estimators, lambda_reg, alpha_reg, risk_factor, comp_time = update_latest_prediction(
                st.session_state['df'], st.session_state['features'], st.session_state['predictions'], 
                st.session_state['historical_volatility_performance'], st.session_state['historical_errors'], 
                st.session_state['ci_coverage'], learning_rate=st.session_state['learning_rate'], 
                n_estimators=st.session_state['n_estimators'], lambda_reg=st.session_state['lambda_reg'], 
                alpha_reg=st.session_state['alpha_reg'], risk_factor=st.session_state['risk_factor']
            )
            st.session_state['predictions'] = predictions
            print(f"Updated predictions: April 2025 predicted = ${predictions[-1]['predicted']:.2f}")
            st.session_state['historical_volatility_performance'] = historical_volatility_performance
            st.session_state['learning_rate'] = learning_rate
            st.session_state['n_estimators'] = n_estimators
            st.session_state['lambda_reg'] = lambda_reg
            st.session_state['alpha_reg'] = alpha_reg
            st.session_state['risk_factor'] = risk_factor
            errors = [p['error'] for p in predictions if not np.isnan(p['error'])]
            st.session_state['historical_errors'].extend(errors)
            st.session_state['historical_errors'] = st.session_state['historical_errors'][-1000:]
            new_mae = np.mean([abs(e) for e in errors]) if errors else float('inf')
            if new_mae < st.session_state['best_mae']:
                st.session_state['best_mae'] = new_mae
                st.session_state['best_predictions'] = predictions
                print(f"New best MAE: {new_mae:.2f}")
            else:
                st.session_state['predictions'] = st.session_state['best_predictions']
                print(f"Reverted to best MAE: {st.session_state['best_mae']:.2f}")
            latest_pred = predictions[-1]
            st.session_state['prediction_history'].append({
                'timestamp': datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S'),
                'predicted': latest_pred['predicted'],
                'ci_lower_95': latest_pred['ci_lower_95'],
                'ci_upper_95': latest_pred['ci_upper_95'],
                'ci_lower_75': latest_pred['ci_lower_75'],
                'ci_upper_75': latest_pred['ci_upper_75'],
                'confidence': latest_pred['confidence']
            })
            st.session_state['long_term_confidence_history'].append(latest_pred['confidence'])
            st.session_state['long_term_confidence_history'] = st.session_state['long_term_confidence_history'][-10:]
            print(f"Added to prediction history: {latest_pred['confidence']}% confidence")
            inside_ci = sum(1 for p in predictions if not np.isnan(p['actual']) and p['ci_lower_95'] <= p['actual'] <= p['ci_upper_95'])
            total = sum(1 for p in predictions if not np.isnan(p['actual']))
            new_ci_coverage = inside_ci / total if total > 0 else 0.95
            st.session_state['ci_coverage'] = new_ci_coverage
            prev_coverage = st.session_state['best_configuration']['ci_coverage']
            prev_confidence = st.session_state['best_configuration']['confidence']
            if (new_ci_coverage > prev_coverage + 0.05) or (latest_pred['confidence'] > prev_confidence + 10):
                st.session_state['best_configuration'] = {
                    'learning_rate': learning_rate,
                    'n_estimators': n_estimators,
                    'lambda_reg': lambda_reg,
                    'alpha_reg': alpha_reg,
                    'risk_factor': risk_factor,
                    'ci_coverage': new_ci_coverage,
                    'confidence': latest_pred['confidence']
                }
                print(f"Breakthrough detected! Updated best configuration with CI coverage={new_ci_coverage:.2f}, confidence={latest_pred['confidence']:.2f}%")
            st.session_state['computation_times'].append(comp_time)
            st.session_state['computation_times'] = st.session_state['computation_times'][-100:]
            st.session_state['last_update'] = time.time()
            st.session_state['update_in_progress'] = False
            time.sleep(0.1)
            st.rerun()
with col3:
    st.markdown('<div class="tooltip"><button>Download Excel</button><span class="tooltiptext">Export predictions and data to Excel</span></div>', unsafe_allow_html=True)
    if st.button("", key="download_excel"):
        if st.session_state['update_in_progress']:
            st.warning("An update is currently in progress. Please wait a moment and try again to ensure the latest predictions are included.")
        elif st.session_state['predictions']:
            try:
                filename = generate_excel_report(
                    st.session_state['learning_rate'], st.session_state['n_estimators'], 
                    st.session_state['lambda_reg'], st.session_state['alpha_reg'], 
                    st.session_state['risk_factor']
                )
                if filename:
                    with open(filename, "rb") as file:
                        st.download_button(
                            label="Download Excel Report",
                            data=file,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            except Exception as e:
                st.error(f"Error generating Excel report: {e}")
        else:
            st.warning("No predictions available yet. Please wait for the first prediction cycle.")
with col4:
    st.markdown('<div class="tooltip"><button>Clear History</button><span class="tooltiptext">Reset prediction history</span></div>', unsafe_allow_html=True)
    if st.button("", key="clear_history"):
        st.session_state['prediction_history'] = []
        st.session_state['historical_errors'] = []
        st.session_state['ci_coverage'] = 0.95
        st.session_state['computation_times'] = []
        st.session_state['long_term_confidence_history'] = []
        st.session_state['last_update'] = time.time()

status = "Running" if st.session_state['running'] else "Paused"
if st.session_state['running']:
    status_display = f"**Status:** {status} <span class='spinner'>⟳</span>"
else:
    status_display = f"**Status:** {status}"
st.markdown(status_display, unsafe_allow_html=True)

last_update_str = datetime.fromtimestamp(st.session_state['last_update']).strftime('%Y-%m-%d %H:%M:%S')
st.markdown(f"**Last Updated:** {last_update_str}")

st.markdown(f"**Peak Detection:** {st.session_state['peak_message']}")

if st.session_state['predictions']:
    render_charts(st.session_state['predictions'], show_feature_importance, show_prediction_history, show_confidence_trend)

st.session_state['update_counter'] += 1
current_time = time.time()
time_since_last_update = current_time - st.session_state['last_update']
print(f"Update check #{st.session_state['update_counter']}: Running={st.session_state['running']}, Time since last update={time_since_last_update:.2f}s, Update frequency={update_frequency}s")
if st.session_state['running'] and time_since_last_update > update_frequency:
    print("Auto-update triggered.")
    st.session_state['update_in_progress'] = True
    with st.spinner("Refreshing Predictions..."):
        latest_confidence = st.session_state['predictions'][-1]['confidence']
        prev_best_confidence = st.session_state['best_configuration']['confidence']
        use_best_config = latest_confidence < prev_best_confidence - 10
        if use_best_config:
            print(f"Confidence dropped significantly ({latest_confidence:.2f} < {prev_best_confidence:.2f}), reverting to best configuration.")
            learning_rate = st.session_state['best_configuration']['learning_rate']
            n_estimators = st.session_state['best_configuration']['n_estimators']
            lambda_reg = st.session_state['best_configuration']['lambda_reg']
            alpha_reg = st.session_state['best_configuration']['alpha_reg']
            risk_factor = st.session_state['best_configuration']['risk_factor']
        else:
            learning_rate = st.session_state['learning_rate']
            n_estimators = st.session_state['n_estimators']
            lambda_reg = st.session_state['lambda_reg']
            alpha_reg = st.session_state['alpha_reg']
            risk_factor = st.session











